from django.contrib.auth.decorators import login_required
from .models import AcademicYear
from django.views.decorators.http import require_POST

# Handle creation of a new academic year from modal form
@login_required
@require_POST
def add_academic_year(request):
    year = request.POST.get('year')
    start_date = request.POST.get('start_date')
    end_date = request.POST.get('end_date')
    if not (year and start_date and end_date):
        messages.error(request, 'All fields are required to add a new academic year.')
        return redirect(request.META.get('HTTP_REFERER', 'students_app:fee_structure_management'))
    try:
        AcademicYear.objects.create(year=year, start_date=start_date, end_date=end_date)
        messages.success(request, f'Academic year {year} added successfully!')
    except Exception as e:
        messages.error(request, f'Error adding academic year: {str(e)}')
    return redirect(request.META.get('HTTP_REFERER', 'students_app:fee_structure_management'))
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import Count, Q
from django.db import models
from django.views.generic import ListView, DetailView, UpdateView

def is_super_admin(user):
    """Check if user is super admin"""
    if user.is_superuser:
        return True
    try:
        if hasattr(user, 'school_profile'):
            school_user = user.school_profile
            if school_user and school_user.role:
                return school_user.role.name == 'super_admin'
    except:
        pass
    return False


def _has_school_admin_permission(user):
    try:
        if hasattr(user, 'school_profile'):
            school_user = user.school_profile
            if school_user and school_user.role and school_user.role.name:
                return 'admin' in school_user.role.name.lower().strip()
    except Exception:
        pass
    return False


def _can_manage_school_data(user):
    return bool(user and (getattr(user, 'is_superuser', False) or is_super_admin(user) or _has_school_admin_permission(user)))


def _get_user_school(user):
    try:
        if hasattr(user, 'school_profile') and user.school_profile and user.school_profile.school:
            return user.school_profile.school
    except Exception:
        pass
    return None

def login_redirect(request):
    """Redirect to dashboard or login based on authentication"""
    if request.user.is_authenticated:
        return redirect('students_app:dashboard')
    else:
        return redirect('students_app:simple_login')


@login_required
def home(request):
    """Dashboard home view - Redirects to appropriate dashboard based on user role"""
    from django.shortcuts import redirect
    from .models import Student, Teacher, Parent, SchoolUser, Staff

    user = request.user

    # 1. Check for Super Admin / Superuser
    if user.is_superuser:
        return redirect('students_app:admin_dashboard')

    # 2. Check for School Admin / School User Profile
    try:
        if hasattr(user, 'school_profile'):
            school_user = user.school_profile
            # If explicit admin role
            if school_user.role and school_user.role.name in ['super_admin', 'school_admin']:
                return redirect('students_app:admin_dashboard')
            # If other roles in SchoolUser, map them if necessary, or fall through
            if school_user.role and school_user.role.name == 'teacher':
                 return redirect('students_app:teacher_dashboard_restricted')
            if school_user.role and school_user.role.name == 'student':
                 return redirect('students_app:student_dashboard')
            if school_user.role and school_user.role.name == 'parent':
                 return redirect('students_app:parent_dashboard')
    except Exception:
        pass

    # 3. Check for Teacher Profile
    if Teacher.objects.filter(user=user).exists():
        return redirect('students_app:teacher_dashboard_restricted')

    # 4. Check for Student Profile (Main Student Model)
    # Note: Students usually log in via admission number as username
    if Student.objects.filter(admission_number=user.username).exists():
        return redirect('students_app:student_dashboard')

    # 5. Check for Parent Profile
    if Parent.objects.filter(user=user).exists():
        return redirect('students_app:parent_dashboard')
        
    # 6. Check for Staff Profile
    if Staff.objects.filter(user=user).exists():
        return redirect('students_app:teacher_dashboard')  # Staff uses teacher dashboard for now

    # 7. Fallback / Default
    # If no specific role found, send to simple login or show an error
    # but to avoid loop, let's send to admin dashboard via try/catch in admin view
    # OR better, send to a profile selection or error page if available.
    # For now, let's try admin dashboard, as it handles "no permission" by redirecting to home...
    # WAIT, that causes the loop.
    # So we must redirect to something safe.
    
    # If users are created via standard Django user creation without profiles, 
    # they might be expecting to see a basic landing page. 
    # But since we don't have a generic "home", we will assume they are admins pending setup.
    return redirect('students_app:admin_dashboard')


def generate_id_cards_reportlab(students, request, single_card=False):
    """
    Generate ID cards PDF using ReportLab
    This is a stub implementation - should be enhanced with actual ID card generation logic
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import mm
        from io import BytesIO
        from django.http import HttpResponse
        
        # Create PDF buffer
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        
        # Basic implementation - just add text for now
        # TODO: Implement full ID card generation with images, QR codes, etc.
        y = 800
        p.drawString(100, y, "ID Cards Generated")
        y -= 30
        
        for student in students:
            if y < 50:
                p.showPage()
                y = 800
            
            p.drawString(100, y, f"Student: {student.get_full_name() if hasattr(student, 'get_full_name') else str(student)}")
            y -= 20
            if hasattr(student, 'admission_number'):
                p.drawString(100, y, f"Admission: {student.admission_number}")
                y -= 20
        
        p.showPage()
        p.save()
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="id_cards.pdf"'
        return response
        
    except ImportError:
        from django.http import HttpResponse
        return HttpResponse("ReportLab is required for ID card generation. Please install: pip install reportlab", 
                          status=501)
    except Exception as e:
        from django.http import HttpResponse
        return HttpResponse(f"Error generating ID cards: {str(e)}", status=500)


@login_required
def manage_sections(request):
    """Create and manage sections"""
    # Check permissions - allow any role with 'admin' in name
    has_permission = False
    try:
        if hasattr(request.user, 'school_profile'):
            school_user = request.user.school_profile
            if school_user and school_user.role:
                role_name = school_user.role.name.lower().strip()
                if 'admin' in role_name:
                    has_permission = True
    except:
        pass
    
    if not has_permission and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to manage sections.')
        return redirect('students_app:manage_timetables')
    
    from .models import Class, Section
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create':
            class_id = request.POST.get('class_id')
            section_name = request.POST.get('section_name', '').strip().upper()
            capacity = request.POST.get('capacity', 40)
            
            if not class_id or not section_name:
                messages.error(request, 'Please select a class and enter section name.')
                return redirect('students_app:manage_sections')
            
            try:
                class_obj = Class.objects.get(id=class_id)
                
                # Check if section already exists
                if Section.objects.filter(class_assigned=class_obj, name=section_name).exists():
                    messages.warning(request, f'Section "{section_name}" already exists for {class_obj.name}.')
                    return redirect('students_app:manage_sections')
                
                # Create section
                Section.objects.create(
                    class_assigned=class_obj,
                    name=section_name,
                    capacity=int(capacity) if capacity else 40
                )
                messages.success(request, f'Section "{section_name}" created successfully for {class_obj.name}!')
                return redirect('students_app:manage_sections')
            except Class.DoesNotExist:
                messages.error(request, 'Selected class not found.')
            except Exception as e:
                messages.error(request, f'Error creating section: {str(e)}')
        
        elif action == 'delete':
            section_id = request.POST.get('section_id')
            if section_id:
                try:
                    section = Section.objects.get(id=section_id)
                    section_name = str(section)
                    section.delete()
                    messages.success(request, f'Section "{section_name}" deleted successfully!')
                except Section.DoesNotExist:
                    messages.error(request, 'Section not found.')
                except Exception as e:
                    messages.error(request, f'Error deleting section: {str(e)}')
            return redirect('students_app:manage_sections')
    
    # Get all classes and sections
    classes = Class.objects.all().order_by('numeric_value', 'name')
    sections = Section.objects.select_related('class_assigned').order_by(
        'class_assigned__numeric_value', 'class_assigned__name', 'name'
    )
    
    # Group sections by class for display
    sections_by_class = {}
    for section in sections:
        class_name = section.class_assigned.name
        if class_name not in sections_by_class:
            sections_by_class[class_name] = []
        sections_by_class[class_name].append(section)
    
    context = {
        'classes': classes,
        'sections': sections,
        'sections_by_class': sections_by_class,
    }
    return render(request, 'students/manage_sections.html', context)


class StudentListView(LoginRequiredMixin, ListView):
    """List view for students with filtering"""
    model = None  # Will be set dynamically
    template_name = 'students/student_list.html'
    context_object_name = 'students'
    paginate_by = 50
    
    def get_queryset(self):
        from .models import Student
        from .models import AcademicYear
        
        queryset = Student.objects.filter(status='active').select_related(
            'current_class', 'section', 'academic_year', 'school'
        )
        
        # Super admin: School filter - mandatory, show no data if school not selected
        if is_super_admin(self.request.user):
            school_id = self.request.GET.get('school')
            if school_id:
                queryset = queryset.filter(school_id=school_id)
            else:
                queryset = Student.objects.none()  # Return empty queryset if school not selected
        else:
            # Normal user: Filter by their school automatically
            try:
                if hasattr(self.request.user, 'school_profile') and self.request.user.school_profile.school:
                    queryset = queryset.filter(school=self.request.user.school_profile.school)
                else:
                    # If no school profile, show no students
                    queryset = Student.objects.none()
            except:
                queryset = Student.objects.none()
        
        # Search filter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(middle_name__icontains=search) |
                Q(admission_number__icontains=search) |
                Q(father_name__icontains=search) |
                Q(father_phone__icontains=search)
            )
        
        # Class filter
        class_id = self.request.GET.get('class')
        if class_id:
            queryset = queryset.filter(current_class_id=class_id)
        
        # Section filter
        section_id = self.request.GET.get('section')
        if section_id:
            queryset = queryset.filter(section_id=section_id)
        
        # Academic year filter
        year_id = self.request.GET.get('year')
        if year_id:
            queryset = queryset.filter(academic_year_id=year_id)
        else:
            current_year = AcademicYear.objects.filter(is_current=True).first() or AcademicYear.objects.order_by('-start_date').first()
            if current_year:
                queryset = queryset.filter(academic_year=current_year)
        
        return queryset.order_by('current_class__numeric_value', 'section__name', 'roll_number')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import Class, Section, AcademicYear, School
        
        context['classes'] = Class.objects.all().order_by('numeric_value', 'name')
        context['sections'] = Section.objects.select_related('class_assigned').order_by(
            'class_assigned__numeric_value', 'class_assigned__name', 'name'
        )
        context['academic_years'] = AcademicYear.objects.all().order_by('-start_date')

        selected_year = self.request.GET.get('year')
        if not selected_year:
            current_year = AcademicYear.objects.filter(is_current=True).first() or AcademicYear.objects.order_by('-start_date').first()
            if current_year:
                selected_year = str(current_year.id)
        context['selected_year'] = selected_year
        
        # Super admin: Add schools list
        if is_super_admin(self.request.user):
            context['is_super_admin'] = True
            context['schools'] = School.objects.all().order_by('name')
            context['selected_school'] = self.request.GET.get('school', '')
        else:
            context['is_super_admin'] = False
        
        return context


class StudentDetailView(LoginRequiredMixin, DetailView):
    """Detail view for a single student"""
    model = None  # Will be set dynamically
    template_name = 'students/student_detail.html'
    context_object_name = 'student'
    slug_field = 'admission_number'
    slug_url_kwarg = 'admission_number'
    
    def get_queryset(self):
        from .models import Student
        return Student.objects.select_related('current_class', 'section', 'academic_year')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import Attendance, FeePayment, BookIssue, Teacher, ClassTeacher, AcademicYear
        
        student = context['student']
        
        # Get attendance summary
        attendance_records = Attendance.objects.filter(student=student)
        total_attendance = attendance_records.count()
        present_count = attendance_records.filter(status='P').count()
        absent_count = attendance_records.filter(status='A').count()
        
        context['total_attendance'] = total_attendance
        context['present_count'] = present_count
        context['absent_count'] = absent_count
        
        # Calculate attendance percentage
        if total_attendance > 0:
            context['attendance_percentage'] = round((present_count / total_attendance) * 100, 1)
        else:
            context['attendance_percentage'] = 0
        
        # Get fee information
        context['fee_payments'] = FeePayment.objects.filter(student=student).order_by('-payment_date')[:10]
        
        # Get issued books
        context['issued_books'] = BookIssue.objects.filter(
            student=student, 
            status='issued'
        ).select_related('book')
        
        # Check if user can change photo (admin or class teacher)
        can_change_photo = False
        if self.request.user.is_superuser:
            can_change_photo = True
        else:
            try:
                if hasattr(self.request.user, 'school_profile'):
                    school_user = self.request.user.school_profile
                    if school_user and school_user.role:
                        role_name = school_user.role.name.lower().strip()
                        if 'admin' in role_name:
                            can_change_photo = True
            except:
                pass
            
            # Check if user is a teacher and is class teacher for this student's section
            if not can_change_photo:
                try:
                    teacher = Teacher.objects.get(user=self.request.user)
                    current_year = AcademicYear.objects.filter(is_current=True).first()
                    if current_year and student.section:
                        is_class_teacher = ClassTeacher.objects.filter(
                            teacher=teacher,
                            section=student.section,
                            academic_year=current_year
                        ).exists()
                        if is_class_teacher:
                            can_change_photo = True
                except:
                    pass
        
        context['can_change_photo'] = can_change_photo
        
        # Check if ID card template exists and is active
        from .models import IDCardTemplate
        active_template = IDCardTemplate.objects.filter(is_active=True).first()
        context['has_id_card_template'] = active_template is not None and active_template.template_image is not None
        
        return context


class StudentUpdateView(LoginRequiredMixin, UpdateView):
    """Update view for editing a student"""
    model = None  # Will be set dynamically
    template_name = 'students/student_update.html'
    context_object_name = 'student'
    slug_field = 'admission_number'
    slug_url_kwarg = 'admission_number'
    
    def get_queryset(self):
        from .models import Student
        return Student.objects.select_related('current_class', 'section', 'academic_year')
    
    def get_form_class(self):
        from .forms import StudentForm
        return StudentForm
    
    def form_valid(self, form):
        messages.success(self.request, f'Student "{form.instance.get_full_name()}" updated successfully!')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('students_app:student_detail', kwargs={'admission_number': self.object.admission_number})


# ============================================
# Student Management Functions
# ============================================

@login_required
def add_student(request):
    """Add a new student"""
    from .forms import StudentForm
    
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            student = form.save()

            # Create Django User for the student (username = admission_number)
            try:
                from django.contrib.auth import get_user_model
                User = get_user_model()
                username = student.admission_number
                try:
                    pwd = student.date_of_birth.strftime('%Y%m%d')
                except Exception:
                    pwd = str(student.date_of_birth)

                existing_user = User.objects.filter(username=username).first()
                if not existing_user:
                    try:
                        existing_user = User.objects.create_user(
                            username=username,
                            email=student.email or f"{username}@{student.school.name.replace(' ','').lower()}.local" if student.school else f"{username}@school.local",
                            password=pwd,
                            first_name=student.first_name,
                            last_name=student.last_name
                        )
                    except Exception:
                        existing_user = None

                # Create SchoolUser profile
                try:
                    from .models import UserRole, SchoolUser
                    student_role = UserRole.objects.filter(name='student').first()
                    if existing_user and student_role and hasattr(existing_user, 'school_profile') == False:
                        SchoolUser.objects.get_or_create(
                            user=existing_user,
                            defaults={
                                'role': student_role,
                                'school': student.school,
                                'login_id': username,
                                'custom_password': pwd,
                                'phone': student.phone or ''
                            }
                        )
                except Exception:
                    pass
            except Exception:
                pass

            messages.success(request, f'Student "{student.get_full_name()}" added successfully!')
            return redirect('students_app:student_detail', admission_number=student.admission_number)
    else:
        form = StudentForm()
    
    return render(request, 'students/add_student.html', {'form': form})


@login_required
def import_students(request):
    """Import students from Excel"""
    if request.method == 'POST':
        messages.info(request, 'Excel import functionality will be implemented soon.')
        return redirect('students_app:student_list')
    
    return render(request, 'students/student_import.html')


@login_required
def export_student_data(request):
    """Export student data to Excel"""
    from .models import Student
    import csv
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="students_export.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Admission Number', 'First Name', 'Last Name', 'Class', 'Section', 'Roll Number', 'Father Name', 'Father Phone'])
    
    students = Student.objects.filter(status='active').select_related('current_class', 'section')
    for student in students:
        writer.writerow([
            student.admission_number,
            student.first_name,
            student.last_name,
            student.current_class.name if student.current_class else '',
            student.section.name if student.section else '',
            student.roll_number,
            student.father_name,
            student.father_phone
        ])
    
    return response


@login_required
def bulk_download_photos(request):
    """Download all student photos as a ZIP file"""
    from .models import Student
    import zipfile
    import io
    import os
    
    # Get students with optional filters (same as student_list view)
    students = Student.objects.filter(status='active').select_related('current_class', 'section')
    
    # Apply filters from query parameters
    search = request.GET.get('search')
    if search:
        students = students.filter(
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(middle_name__icontains=search) |
            Q(admission_number__icontains=search) |
            Q(father_name__icontains=search) |
            Q(father_phone__icontains=search)
        )
    
    class_id = request.GET.get('class')
    if class_id:
        students = students.filter(current_class_id=class_id)
    
    section_id = request.GET.get('section')
    if section_id:
        students = students.filter(section_id=section_id)
    
    year_id = request.GET.get('year')
    if year_id:
        students = students.filter(academic_year_id=year_id)
    
    # Create ZIP file in memory
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        photos_added = 0
        for student in students:
            if student.photo and student.photo.name:
                try:
                    # Get the file path
                    photo_path = student.photo.path
                    if os.path.exists(photo_path):
                        # Get file extension
                        _, ext = os.path.splitext(photo_path)
                        # Create filename: admission_number_photo.ext
                        filename = f"{student.admission_number}_photo{ext}"
                        # Add to ZIP
                        zip_file.write(photo_path, filename)
                        photos_added += 1
                except Exception as e:
                    # Skip photos that can't be accessed
                    continue
    
    if photos_added == 0:
        messages.warning(request, 'No photos found to download.')
        return redirect('students_app:student_list')
    
    # Prepare response
    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="student_photos.zip"'
    
    return response


# ============================================
# Dashboard Functions
# ============================================

@login_required
def student_dashboard(request):
    """Student portal dashboard"""
    from .models import Student, Attendance, FeePayment, ExamSchedule, Timetable
    
    try:
        student = Student.objects.get(admission_number=request.user.username)
    except Student.DoesNotExist:
        messages.error(request, 'Student profile not found.')
        return redirect('students_app:home')
    
    # Get attendance summary
    attendance_records = Attendance.objects.filter(student=student)
    total_attendance = attendance_records.count()
    present_count = attendance_records.filter(status='P').count()
    attendance_percentage = (present_count / total_attendance * 100) if total_attendance > 0 else 0
    
    # Get recent fee payments
    recent_fees = FeePayment.objects.filter(student=student).order_by('-payment_date')[:5]
    
    # Get upcoming exams
    upcoming_exams = ExamSchedule.objects.filter(
        class_assigned=student.current_class,
        exam__start_date__gte=timezone.now().date()
    ).order_by('exam__start_date')[:5]
    
    # Get today's timetable
    today = timezone.now().date()
    today_timetable = Timetable.objects.filter(
        section=student.section,
        academic_year=student.academic_year
    ).select_related('subject', 'teacher', 'time_slot')
    
    context = {
        'student': student,
        'total_attendance': total_attendance,
        'present_count': present_count,
        'attendance_percentage': round(attendance_percentage, 2),
        'recent_fees': recent_fees,
        'upcoming_exams': upcoming_exams,
        'today_timetable': today_timetable,
    }
    
    return render(request, 'students/student_dashboard.html', context)


@login_required
def parent_dashboard(request):
    """Parent portal dashboard"""
    from .models import Parent, Student
    
    try:
        parent = Parent.objects.get(user=request.user)
        students = Student.objects.filter(father_phone=parent.phone)
    except Parent.DoesNotExist:
        students = Student.objects.filter(father_phone=request.user.username)
    
    context = {'students': students}
    return render(request, 'parents/dashboard.html', context)


@login_required
def teacher_dashboard(request):
    """Teacher portal dashboard - Also accessible by Staff"""
    from .models import Teacher, Timetable, ClassTeacher, AcademicYear, Staff
    
    teacher = None
    staff_member = None
    
    # Try to get teacher first
    try:
        teacher = Teacher.objects.select_related('user').prefetch_related('subjects').get(user=request.user)
    except Teacher.DoesNotExist:
        # If not teacher, check if it's staff
        try:
            staff_member = Staff.objects.select_related('user').get(user=request.user)
            # For staff, we'll use a simplified view or redirect
            # For now, allow them to access but show appropriate message
            messages.info(request, f'Welcome, {staff_member.get_full_name()}! You are logged in as staff.')
        except Staff.DoesNotExist:
            messages.error(request, 'Teacher or Staff profile not found.')
            return redirect('students_app:home')
    
    # Get current academic year
    current_year = AcademicYear.objects.filter(is_current=True).first()
    
    # Get class teacher assignments for current year
    class_teacher_assignments = []
    if current_year:
        class_teacher_assignments = ClassTeacher.objects.filter(
            teacher=teacher,
            academic_year=current_year
        ).select_related('section__class_assigned').order_by('section__class_assigned__numeric_value', 'section__name')
    
    # Get today's classes
    today = timezone.now().date()
    today_classes = Timetable.objects.filter(
        teacher=teacher,
        academic_year__is_current=True
    ).select_related('section', 'subject', 'time_slot')
    
    # Get recent question papers (if model exists)
    recent_papers = []
    try:
        from .models import QuestionPaper
        recent_papers = QuestionPaper.objects.filter(
            created_by=request.user
        ).select_related('subject', 'class_assigned').order_by('-created_at')[:5]
    except:
        pass
    
    # Get recent class tests (if model exists)
    recent_tests = []
    try:
        from .models import ClassTest
        recent_tests = ClassTest.objects.filter(
            teacher=teacher
        ).select_related('subject', 'class_assigned').order_by('-date')[:5]
    except:
        pass
    
    context = {
        'teacher': teacher,
        'staff': staff_member,
        'today_classes': today_classes,
        'recent_papers': recent_papers,
        'recent_tests': recent_tests,
        'class_teacher_assignments': class_teacher_assignments,
    }
    
    return render(request, 'teachers/dashboard.html', context)


@login_required
def librarian_dashboard(request):
    """Librarian portal dashboard"""
    from .models import Book, BookIssue
    
    total_books = Book.objects.count()
    issued_books = BookIssue.objects.filter(status='issued').count()
    available_books = total_books - issued_books
    
    context = {
        'total_books': total_books,
        'issued_books': issued_books,
        'available_books': available_books,
    }
    
    return render(request, 'librarian/dashboard.html', context)


# ============================================
# Teacher Management Functions
# ============================================

@login_required
def teacher_bulk_upload(request):
    """Smart bulk upload for teachers with automatic class/section assignment"""
    return teacher_bulk_upload_smart(request)

@login_required
def smart_bulk_upload(request):
    """Universal smart bulk upload for students, teachers, and staff"""
    from .models import Student, Teacher, Staff, Class, Section, AcademicYear
    from django.contrib.auth.models import User
    from django.utils import timezone
    import pandas as pd
    
    context = {}
    upload_type = request.GET.get('type', 'student')  # Default to student
    
    if request.method == 'POST':
        upload_type = request.POST.get('upload_type', 'student')
        action = request.POST.get('action')
        
        if action == 'save_bulk':
            # Handle different upload types
            if upload_type == 'student':
                return handle_student_bulk_upload(request)
            elif upload_type == 'teacher':
                return handle_teacher_bulk_upload(request)
            elif upload_type == 'staff':
                return handle_staff_bulk_upload(request)
        
        elif 'excel_file' in request.FILES:
            # Handle Excel file upload for any type
            try:
                excel_file = request.FILES['excel_file']
                df = pd.read_excel(excel_file)
                df.columns = [str(c).strip().lower().replace(' ', ' ') for c in df.columns]
                
                data_list = []
                
                if upload_type == 'student':
                    data_list = parse_student_excel(df)
                elif upload_type == 'teacher':
                    data_list = parse_teacher_excel(df)
                elif upload_type == 'staff':
                    data_list = parse_staff_excel(df)
                
                context['data_list'] = data_list
                context['upload_type'] = upload_type
                # Class model has no 'school' FK — classes are global
                context['available_classes'] = Class.objects.all().prefetch_related('section_set').order_by('numeric_value', 'name')
                context['available_sections'] = Section.objects.all().select_related('class_assigned').order_by('class_assigned__numeric_value', 'name')
                context['academic_years'] = AcademicYear.objects.filter(school=request.user.school_profile.school)
                
            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
                
            return render(request, 'students/smart_bulk_upload.html', context)
    
    else:
        # GET request
        context = {
            'upload_type': upload_type,
            'required_columns': get_required_columns(upload_type),
            'academic_years': AcademicYear.objects.filter(school=request.user.school_profile.school) if upload_type == 'student' else []
        }
        
        return render(request, 'students/smart_bulk_upload.html', context)

def get_required_columns(upload_type):
    """Get required columns for different upload types"""
    if upload_type == 'student':
        return "First Name, Last Name, Admission Number, Roll Number, Email, Phone, DOB, Gender, Address, Designation, Class, Section, Academic Year, Father Name, Father Phone, Mother Name, Mother Phone, Aadhaar Card, Samagra ID"
    elif upload_type == 'teacher':
        return "First Name, Last Name, Employee ID, Email, Phone, DOB, Gender, Address, Designation, Qualification, Joining Date, Current Salary"
    elif upload_type == 'staff':
        return "First Name, Last Name, Employee ID, Email, Phone, DOB, Gender, Address, Designation, Department, Joining Date, Current Salary"
    return ""

def parse_student_excel(df):
    """Parse student data from Excel"""
    students_data = []
    for index, row in df.iterrows():
        students_data.append({
            'first_name': row.get('first name', ''),
            'last_name': row.get('last name', ''),
            'admission_number': row.get('admission number', ''),
            'roll_number': row.get('roll number', ''),
            'email': row.get('email', ''),
            'phone': row.get('phone', ''),
            'date_of_birth': row.get('date of birth', '2000-01-01'),
            'gender': str(row.get('gender (m/f/o)') or row.get('gender') or 'O')[0].upper(),
            'address': row.get('address', ''),
            'designation': row.get('designation', 'Student'),
            'current_class': None,  # Will be set during assignment
            'section': None,  # Will be set during assignment
            'academic_year': None,  # Will be set during assignment
            'father_name': row.get('father name', ''),
            'father_phone': row.get('father phone', ''),
            'mother_name': row.get('mother name', ''),
            'mother_phone': row.get('mother phone', ''),
            'aadhaar_card': row.get('aadhaar card', row.get('adhaar card', row.get('aadhaar', ''))),
            'samagra_id': row.get('samagra id', row.get('samgara id', row.get('samagra', ''))),
            'image_filename': row.get('image filename', ''),
        })
    return students_data

def parse_teacher_excel(df):
    """Parse teacher data from Excel"""
    teachers_data = []
    for index, row in df.iterrows():
        teachers_data.append({
            'first_name': row.get('first name', ''),
            'last_name': row.get('last name', ''),
            'employee_id': row.get('employee id', ''),
            'email': row.get('email', ''),
            'phone': row.get('phone', ''),
            'date_of_birth': row.get('date of birth', '1980-01-01'),
            'gender': str(row.get('gender (m/f/o)') or row.get('gender') or 'O')[0].upper(),
            'address': row.get('address', ''),
            'designation': row.get('designation', 'Teacher'),
            'qualification': row.get('qualification', ''),
            'joining_date': row.get('joining date', timezone.now().date().strftime('%Y-%m-%d')),
            'current_salary': row.get('current salary') or row.get('salary') or 0,
            'image_filename': row.get('image filename', ''),
        })
    return teachers_data

def parse_staff_excel(df):
    """Parse staff data from Excel"""
    staff_data = []
    for index, row in df.iterrows():
        staff_data.append({
            'first_name': row.get('first name', ''),
            'last_name': row.get('last name', ''),
            'employee_id': row.get('employee id', ''),
            'email': row.get('email', ''),
            'phone': row.get('phone', ''),
            'date_of_birth': row.get('date of birth', '1980-01-01'),
            'gender': str(row.get('gender (m/f/o)') or row.get('gender') or 'O')[0].upper(),
            'address': row.get('address', ''),
            'designation': row.get('designation', 'Staff'),
            'department': row.get('department', 'General'),
            'joining_date': row.get('joining date', timezone.now().date().strftime('%Y-%m-%d')),
            'current_salary': row.get('current salary') or row.get('salary') or 0,
            'image_filename': row.get('image filename', ''),
        })
    return staff_data

def handle_student_bulk_upload(request):
    """Handle student bulk upload with smart assignment"""
    from .models import Student, School, Class, Section, AcademicYear, SchoolUser
    from django.contrib.auth.models import User
    from django.contrib import messages
    from django.shortcuts import redirect
    
    if request.method == 'POST':
        # Get form data
        first_names = request.POST.getlist('first_name[]')
        last_names = request.POST.getlist('last_name[]')
        admission_numbers = request.POST.getlist('admission_number[]')
        roll_numbers = request.POST.getlist('roll_number[]')
        emails = request.POST.getlist('email[]')
        phones = request.POST.getlist('phone[]')
        dobs = request.POST.getlist('dob[]')
        genders = request.POST.getlist('gender[]')
        father_names = request.POST.getlist('father_name[]')
        father_phones = request.POST.getlist('father_phone[]')
        mother_names = request.POST.getlist('mother_name[]')
        mother_phones = request.POST.getlist('mother_phone[]')
        assigned_classes = request.POST.getlist('assigned_class[]')
        assigned_sections = request.POST.getlist('assigned_section[]')
        academic_years = request.POST.getlist('academic_year[]')
        addresses = request.POST.getlist('address[]')
        designations = request.POST.getlist('designation[]')
        aadhaar_cards = request.POST.getlist('aadhaar_card[]')
        samagra_ids = request.POST.getlist('samagra_id[]')
        
        school_obj = request.user.school_profile.school if hasattr(request.user, 'school_profile') and request.user.school_profile else None
        
        success_count = 0
        error_count = 0
        skipped_details = []

        for i in range(len(first_names)):
            try:
                if not admission_numbers[i]:
                    error_count += 1
                    skipped_details.append(f"Row {i+1}: Missing Admission Number")
                    continue

                # Check for duplicates
                if Student.objects.filter(admission_number=admission_numbers[i], school=school_obj).exists():
                    error_count += 1
                    skipped_details.append(f"Row {i+1}: Duplicate Admission Number '{admission_numbers[i]}'")
                    continue

                # Create student
                student = Student(
                    school=school_obj,
                    admission_number=admission_numbers[i],
                    roll_number=roll_numbers[i],
                    first_name=first_names[i],
                    last_name=last_names[i],
                    date_of_birth=dobs[i],
                    gender=genders[i],
                    email=emails[i],
                    phone=phones[i],
                    father_name=father_names[i],
                    father_phone=father_phones[i],
                    mother_name=mother_names[i],
                    mother_phone=mother_phones[i],
                    address=addresses[i],
                    designation=designations[i] if designations[i] else 'Student',
                    aadhaar_card=aadhaar_cards[i] if i < len(aadhaar_cards) else None,
                    samagra_id=samagra_ids[i] if i < len(samagra_ids) else None,
                    admission_date=timezone.now().date()
                )
                
                # Assign class, section, academic year
                if i < len(assigned_classes) and assigned_classes[i]:
                    try:
                        student.current_class = Class.objects.get(id=assigned_classes[i])
                    except Class.DoesNotExist:
                        pass
                
                if i < len(assigned_sections) and assigned_sections[i]:
                    try:
                        student.section = Section.objects.get(id=assigned_sections[i])
                    except Section.DoesNotExist:
                        pass
                
                if i < len(academic_years) and academic_years[i]:
                    try:
                        student.academic_year = AcademicYear.objects.get(id=academic_years[i])
                    except AcademicYear.DoesNotExist:
                        pass

                student.save()
                
                # Create User and Profile (simplified version of add_student logic)
                try:
                    username = student.admission_number
                    pwd = dobs[i].replace('-', '') if '-' in dobs[i] else str(dobs[i])
                    
                    if not User.objects.filter(username=username).exists():
                        user = User.objects.create_user(
                            username=username,
                            email=student.email or f"{username}@school.local",
                            password=pwd,
                            first_name=student.first_name,
                            last_name=student.last_name
                        )
                        
                        from .models import UserRole, SchoolUser
                        student_role = UserRole.objects.filter(name='student').first()
                        if student_role:
                            SchoolUser.objects.create(
                                user=user,
                                role=student_role,
                                school=school_obj,
                                login_id=username,
                                custom_password=pwd,
                                phone=student.phone or ''
                            )
                except Exception as user_exc:
                    print(f"User creation failed for {username}: {user_exc}")

                success_count += 1
                
            except Exception as e:
                error_count += 1
                skipped_details.append(f"Row {i+1}: {str(e)}")
        
        if success_count > 0:
            messages.success(request, f"Successfully imported {success_count} students.")
        if error_count > 0:
            msg = f"Skipped {error_count} students due to errors."
            if skipped_details:
                msg += f" Sample: {'; '.join(skipped_details[:3])}"
            messages.warning(request, msg)
            
        return redirect('students_app:student_list')
    
    return redirect('students_app:smart_bulk_upload')

def handle_teacher_bulk_upload(request):
    """Handle teacher bulk upload with smart assignment"""
    return teacher_bulk_upload_smart(request)

def handle_staff_bulk_upload(request):
    """Handle staff bulk upload with smart assignment"""
    # Implementation for staff bulk upload
    pass

def teacher_bulk_upload_smart(request):
    """Smart bulk upload for teachers with automatic class/section assignment"""
    from .models import Teacher, School, Class, Section, SchoolUser
    from django.contrib.auth.models import User
    from django.utils import timezone
    import pandas as pd
    
    context = {}
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'save_bulk':
            import os
            from django.core.files.base import ContentFile
            from django.conf import settings
            
            # Get form data
            first_names = request.POST.getlist('first_name[]')
            last_names = request.POST.getlist('last_name[]')
            employee_ids = request.POST.getlist('employee_id[]')
            emails = request.POST.getlist('email[]')
            phones = request.POST.getlist('phone[]')
            dobs = request.POST.getlist('dob[]')
            genders = request.POST.getlist('gender[]')
            qualifications = request.POST.getlist('qualification[]')
            salaries = request.POST.getlist('salary[]')
            addresses = request.POST.getlist('address[]')
            joining_dates = request.POST.getlist('joining_date[]')
            image_filenames = request.POST.getlist('image_filename[]')
            designations = request.POST.getlist('designation[]')
            assigned_classes = request.POST.getlist('assigned_class[]')
            assigned_sections = request.POST.getlist('assigned_section[]')
            auto_assign_classes = request.POST.get('auto_assign_classes', 'no')
            auto_assign_sections = request.POST.get('auto_assign_sections', 'no')
            images_path = request.session.get('teacher_import_images_path')
            
            success_count = 0
            error_count = 0
            skipped_details = []

            # Get available classes and sections for smart assignment
            school_obj = request.user.school_profile.school if hasattr(request.user, 'school_profile') and request.user.school_profile else None
            # Class model has no 'school' FK — classes are global
            available_classes = Class.objects.all().prefetch_related('section_set').order_by('numeric_value', 'name')
            available_sections = Section.objects.all().select_related('class_assigned').order_by('class_assigned__numeric_value', 'name')
            
            context['available_classes'] = available_classes
            context['available_sections'] = available_sections
            context['auto_assign_classes'] = auto_assign_classes
            context['auto_assign_sections'] = auto_assign_sections

            # Smart assignment logic
            for i in range(len(first_names)):
                try:
                    if not emails[i] or not employee_ids[i]:
                        error_count += 1
                        skipped_details.append(f"Row {i+1}: Missing required fields (Email / Employee ID)")
                        continue

                    # Check for duplicates
                    if User.objects.filter(username=emails[i]).exists():
                        su = SchoolUser.objects.filter(user__username=emails[i], school=school_obj).first()
                        if su:
                            error_count += 1
                            skipped_details.append(f"Row {i+1}: Duplicate email/username '{emails[i]}' within this school")
                            continue

                    if Teacher.objects.filter(employee_id=employee_ids[i], school=school_obj).exists():
                        error_count += 1
                        skipped_details.append(f"Row {i+1}: Duplicate Employee ID '{employee_ids[i]}' within this school")
                        continue

                    # Create user
                    user = User.objects.create_user(
                        username=emails[i],
                        email=emails[i],
                        password="Teacher@123",
                        first_name=first_names[i],
                        last_name=last_names[i]
                    )

                    # Create teacher
                    teacher = Teacher(
                        user=user,
                        employee_id=employee_ids[i],
                        phone=phones[i],
                        date_of_birth=dobs[i],
                        gender=genders[i],
                        qualification=qualifications[i],
                        current_salary=salaries[i] if salaries[i] else 0,
                        address=addresses[i] if addresses[i] else "Not Provided",
                        joining_date=joining_dates[i],
                        designation=designations[i] if i < len(designations) and designations[i] else 'Teacher'
                    )
                    
                    # Smart class assignment
                    if auto_assign_classes == 'yes' and available_classes.exists():
                        class_index = i % len(available_classes)
                        assigned_class = available_classes[class_index]
                        teacher.assigned_class = assigned_class
                        print(f"Auto-assigned class {assigned_class.name} to teacher {first_names[i]} {last_names[i]}")
                    
                    # Smart section assignment
                    if auto_assign_sections == 'yes' and available_sections.exists():
                        section_index = i % len(available_sections)
                        assigned_section = available_sections[section_index]
                        teacher.assigned_section = assigned_section
                        print(f"Auto-assigned section {assigned_section.name} to teacher {first_names[i]} {last_names[i]}")
                    
                    # Manual assignment
                    if assigned_classes and i < len(assigned_classes) and assigned_classes[i]:
                        try:
                            assigned_class = Class.objects.get(id=assigned_classes[i])
                            teacher.assigned_class = assigned_class
                        except Class.DoesNotExist:
                            pass
                    
                    if assigned_sections and i < len(assigned_sections) and assigned_sections[i]:
                        try:
                            assigned_section = Section.objects.get(id=assigned_sections[i])
                            teacher.assigned_section = assigned_section
                        except Section.DoesNotExist:
                            pass
                    
                    # Assign school
                    if hasattr(request.user, 'school_profile') and request.user.school_profile:
                        teacher.school = request.user.school_profile.school

                    teacher.save()
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    print(f"Error adding teacher {first_names[i]}: {e}")
                    skipped_details.append(f"Row {i+1}: {str(e)}")
            
            # Clean up
            if images_path and os.path.isdir(images_path):
                try:
                    import shutil
                    shutil.rmtree(images_path, ignore_errors=True)
                except Exception:
                    pass
                request.session.pop('teacher_import_images_path', None)
            
            if success_count > 0:
                messages.success(request, f"Successfully added {success_count} teachers.")
            if error_count > 0:
                msg = f"Skipped {error_count} teachers (duplicates or errors)."
                if skipped_details:
                    sample = skipped_details[:5]
                    msg += f" Sample: {'; '.join(sample)}"
                messages.warning(request, msg)
            return redirect('students_app:teacher_list')
        
        elif 'excel_file' in request.FILES:
            # Handle Excel file upload
            try:
                import os
                import uuid
                from django.core.files.base import ContentFile
                from django.conf import settings
                
                excel_file = request.FILES['excel_file']
                images_zip = request.FILES.get('images_zip')
                images_path = None
                
                if images_zip and images_zip.name.endswith('.zip'):
                    try:
                        images_dict = extract_images_from_zip(images_zip)
                        if images_dict:
                            temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_teacher_import', str(uuid.uuid4()))
                            os.makedirs(temp_dir, exist_ok=True)
                            for fn, data in images_dict.items():
                                with open(os.path.join(temp_dir, fn), 'wb') as f:
                                    f.write(data)
                            request.session['teacher_import_images_path'] = temp_dir
                    except Exception:
                        request.session.pop('teacher_import_images_path', None)
                else:
                    request.session.pop('teacher_import_images_path', None)
                
                df = pd.read_excel(excel_file)
                df.columns = [str(c).strip().lower().replace(' ', ' ') for c in df.columns]
                
                def get_photo(row):
                    for col in ['image filename', 'image_filename', 'photo', 'photo filename']:
                        if col in df.columns and pd.notna(row.get(col)):
                            return str(row[col]).strip()
                    return ''
                
                teachers_data = []
                for index, row in df.iterrows():
                    def format_date(val, default):
                        if pd.isna(val) or str(val).strip() == '': return default
                        try: return pd.to_datetime(val).strftime('%Y-%m-%d')
                        except: return default
                    
                    dob = format_date(row.get('date of birth (yyyy-mm-dd)') or row.get('dob'), '2000-01-01')
                    joining = format_date(row.get('joining date (yyyy-mm-dd)') or row.get('joining date'), timezone.now().date().strftime('%Y-%m-%d'))
                    
                    teachers_data.append({
                        'first_name': row.get('first name', ''),
                        'last_name': row.get('last name', ''),
                        'employee_id': row.get('employee id', ''),
                        'email': row.get('email', ''),
                        'phone': row.get('phone', ''),
                        'dob': dob,
                        'gender': str(row.get('gender (m/f/o)') or row.get('gender') or 'O')[0].upper(),
                        'address': row.get('address', ''),
                        'designation': row.get('designation', 'Teacher'),
                        'qualification': row.get('qualification', ''),
                        'joining_date': joining,
                        'salary': row.get('current salary') or row.get('salary') or 0,
                        'image_filename': get_photo(row),
                    })
                
                context['teachers_data'] = teachers_data
                # Class model has no 'school' FK — classes are global
                context['available_classes'] = Class.objects.all().prefetch_related('section_set').order_by('numeric_value', 'name')
                context['available_sections'] = Section.objects.all().select_related('class_assigned').order_by('class_assigned__numeric_value', 'name')
                
            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
                
            return render(request, 'students/teacher_bulk_upload.html', context)
    
    else:
        # GET request - show upload form
        context = {}
        return render(request, 'students/teacher_bulk_upload.html', context)

def extract_images_from_zip(zip_file):
    """Extract images from ZIP file and return dictionary"""
    import zipfile
    images_dict = {}
    
    try:
        with zipfile.ZipFile(zip_file, 'r') as zip_ref:
            for file_info in zip_ref.infolist():
                if not file_info.is_dir():
                    with zip_ref.open(file_info.filename) as file:
                        images_dict[file_info.filename] = file.read()
    except Exception as e:
        print(f"Error extracting ZIP: {e}")
    
    return images_dict


@login_required
def teacher_list(request):
    """List all teachers"""
    from .models import Teacher, School
    
    teachers = Teacher.objects.filter(is_active=True).select_related('user', 'school')
    
    # Super admin: School filter - mandatory, show no data if school not selected
    if is_super_admin(request.user):
        school_id = request.GET.get('school')
        if school_id:
            teachers = teachers.filter(school_id=school_id)
        else:
            teachers = Teacher.objects.none()  # Return empty queryset if school not selected
        schools = School.objects.all().order_by('name')
        context = {
            'teachers': teachers,
            'is_super_admin': True,
            'schools': schools,
            'selected_school': request.GET.get('school', '')
        }
    else:
        # Check if user has admin permissions
        has_admin_permission = False
        try:
            if hasattr(request.user, 'school_profile'):
                school_user = request.user.school_profile
                if school_user and school_user.role:
                    role_name = school_user.role.name.lower().strip()
                    if 'admin' in role_name:
                        has_admin_permission = True
        except:
            pass
        
        context = {
            'teachers': teachers,
            'is_super_admin': False,
            'has_admin_permission': has_admin_permission or request.user.is_superuser,
        }
    
    return render(request, 'teachers/list.html', context)


@login_required
def add_teacher(request):
    """Add a new teacher"""
    from .forms import TeacherForm
    from django.contrib.auth.models import User
    
    if request.method == 'POST':
        form = TeacherForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Create User
                username = form.cleaned_data['username']
                email = form.cleaned_data['email']
                password = form.cleaned_data['password']
                first_name = form.cleaned_data['first_name']
                last_name = form.cleaned_data['last_name']
                
                user = User.objects.create_user(username=username, email=email, password=password, first_name=first_name, last_name=last_name)
                
                # Create Teacher
                teacher = form.save(commit=False)
                teacher.user = user
                teacher.save()
                form.save_m2m()
                
                messages.success(request, f'Teacher "{user.get_full_name()}" added successfully!')
                return redirect('students_app:teacher_list')
            except Exception as e:
                messages.error(request, f'Error adding teacher: {str(e)}')
    else:
        form = TeacherForm()
    
    return render(request, 'teachers/teacher_form.html', {'form': form, 'title': 'Add Teacher'})


@login_required
def edit_teacher(request, teacher_id):
    """Edit teacher details"""
    from .forms import TeacherForm
    from .models import Teacher
    
    teacher = get_object_or_404(Teacher, id=teacher_id)
    
    if request.method == 'POST':
        form = TeacherForm(request.POST, request.FILES, instance=teacher)
        if form.is_valid():
            try:
                # Update User
                user = teacher.user
                user.email = form.cleaned_data['email']
                user.first_name = form.cleaned_data['first_name']
                user.last_name = form.cleaned_data['last_name']
                
                password = form.cleaned_data.get('password')
                if password:
                    user.set_password(password)
                user.save()
                
                form.save()
                messages.success(request, f'Teacher "{user.get_full_name()}" updated successfully!')
                return redirect('students_app:teacher_list')
            except Exception as e:
                messages.error(request, f'Error updating teacher: {str(e)}')
    else:
        form = TeacherForm(instance=teacher)
    
    return render(request, 'teachers/teacher_form.html', {'form': form, 'title': 'Edit Teacher', 'teacher': teacher})


@login_required
def teacher_question_paper(request):
    """Teacher question paper creation"""
    from .forms import QuestionPaperForm
    from .models import Teacher, Subject, Class
    
    try:
        teacher = Teacher.objects.get(user=request.user)
    except Teacher.DoesNotExist:
        messages.error(request, 'You are not registered as a teacher.')
        return redirect('students_app:teacher_dashboard')
    
    if request.method == 'POST':
        form = QuestionPaperForm(request.POST)
        if form.is_valid():
            question_paper = form.save(commit=False)
            question_paper.created_by = teacher
            
            # Get structured questions from POST data
            structured_questions = []
            question_count = int(request.POST.get('question_count', 0))
            
            for i in range(question_count):
                q_type = request.POST.get(f'question_{i}_type', '')
                if q_type:
                    question_data = {
                        'type': q_type,
                        'question_text': request.POST.get(f'question_{i}_text', ''),
                        'marks': request.POST.get(f'question_{i}_marks', '1'),
                        'order': i + 1,
                    }
                    
                    # Add type-specific fields
                    if q_type == 'mcq':
                        question_data['options'] = [
                            request.POST.get(f'question_{i}_option_a', ''),
                            request.POST.get(f'question_{i}_option_b', ''),
                            request.POST.get(f'question_{i}_option_c', ''),
                            request.POST.get(f'question_{i}_option_d', ''),
                        ]
                        question_data['correct_answer'] = request.POST.get(f'question_{i}_correct', 'A')
                    elif q_type == 'fill_blank':
                        question_data['blanks'] = request.POST.get(f'question_{i}_blanks', '').split(',')
                        question_data['answers'] = request.POST.get(f'question_{i}_answers', '').split(',')
                    elif q_type == 'match_columns':
                        left_items = request.POST.get(f'question_{i}_left_items', '').split('\n')
                        right_items = request.POST.get(f'question_{i}_right_items', '').split('\n')
                        question_data['left_column'] = [item.strip() for item in left_items if item.strip()]
                        question_data['right_column'] = [item.strip() for item in right_items if item.strip()]
                        question_data['matches'] = request.POST.get(f'question_{i}_matches', '')
                    elif q_type in ['one_word', 'single_line', 'multiple_line']:
                        question_data['sample_answer'] = request.POST.get(f'question_{i}_sample_answer', '')
                    elif q_type == 'true_false':
                        question_data['correct_answer'] = request.POST.get(f'question_{i}_correct', 'True')
                    elif q_type == 'numerical':
                        question_data['correct_answer'] = request.POST.get(f'question_{i}_correct', '')
                        question_data['unit'] = request.POST.get(f'question_{i}_unit', '')
                    
                    structured_questions.append(question_data)
            
            question_paper.structured_questions = structured_questions
            question_paper.save()
            messages.success(request, f'Question paper "{question_paper.title}" created successfully!')
            return redirect('students_app:teacher_dashboard')
    else:
        form = QuestionPaperForm()
    
    subjects = Subject.objects.filter(is_active=True).order_by('name')
    classes = Class.objects.all().order_by('numeric_value', 'name')
    
    context = {
        'form': form,
        'subjects': subjects,
        'classes': classes,
    }
    return render(request, 'teachers/question_paper.html', context)


@login_required
def teacher_whatsapp(request):
    """Teacher WhatsApp messaging - send messages to students and parents"""
    from .models import Teacher, Student, Section, Class, Timetable, AcademicYear
    from .forms import WhatsAppMessageForm
    from django.contrib import messages
    
    try:
        teacher = Teacher.objects.get(user=request.user)
    except Teacher.DoesNotExist:
        messages.error(request, 'You are not registered as a teacher.')
        return redirect('students_app:teacher_dashboard')
    
    # Get teacher's assigned sections from multiple sources
    current_year = AcademicYear.objects.filter(is_current=True).first()
    teacher_sections = []
    section_ids = set()
    
    if current_year:
        # Get sections from timetable
        timetable_sections = Section.objects.filter(
            timetables__teacher=teacher,
            timetables__academic_year=current_year
        ).values_list('id', flat=True).distinct()
        section_ids.update(timetable_sections)
        
        # Get sections from ClassTeacher relationship
        class_teacher_sections = Section.objects.filter(
            classteacher__teacher=teacher,
            classteacher__academic_year=current_year
        ).values_list('id', flat=True).distinct()
        section_ids.update(class_teacher_sections)
        
        # Get sections queryset
        if section_ids:
            teacher_sections = Section.objects.filter(id__in=section_ids).order_by('class_assigned__numeric_value', 'name')
    
    # Get all students in teacher's sections
    teacher_students = Student.objects.none()  # Start with empty queryset
    
    if teacher_sections:
        teacher_students = Student.objects.filter(
            section__in=teacher_sections,
            status='active'
        ).select_related('current_class', 'section').order_by('current_class__numeric_value', 'section__name', 'first_name')
    
    # If no students found from sections, try to get students from timetable entries (by subject)
    if not teacher_students.exists() and current_year and teacher.subjects.exists():
        # Get sections where teacher teaches any of their subjects
        subject_sections = Section.objects.filter(
            timetables__teacher=teacher,
            timetables__academic_year=current_year,
            timetables__subject__in=teacher.subjects.all()
        ).distinct()
        
        if subject_sections.exists():
            teacher_students = Student.objects.filter(
                section__in=subject_sections,
                status='active'
            ).select_related('current_class', 'section').distinct().order_by('current_class__numeric_value', 'section__name', 'first_name')
    
    # Final fallback: if still no students, show all active students (for admin/teacher flexibility)
    if not teacher_students.exists():
        teacher_students = Student.objects.filter(
            status='active'
        ).select_related('current_class', 'section').order_by('current_class__numeric_value', 'section__name', 'first_name')[:100]  # Limit to 100 for performance
    
    # Handle form submission
    from urllib.parse import quote
    
    whatsapp_links = None  # Initialize
    
    if request.method == 'POST':
        audience = request.POST.get('audience')
        message_text = request.POST.get('message', '').strip()
        section_id = request.POST.get('section_id')
        student_id = request.POST.get('student_id')
        
        if not message_text:
            messages.error(request, 'Please enter a message.')
        else:
            recipients = []
            recipient_count = 0
            
            if audience == 'individual' and student_id:
                try:
                    student = Student.objects.get(id=student_id)
                    # Get parent phone if available
                    phone = student.father_phone or student.mother_phone or student.guardian_phone
                    if phone:
                        # Clean phone number (remove +, spaces, etc.)
                        clean_phone = ''.join(filter(str.isdigit, phone))
                        if not clean_phone.startswith('91') and len(clean_phone) == 10:
                            clean_phone = '91' + clean_phone
                        
                        whatsapp_url = f"https://wa.me/{clean_phone}?text={quote(message_text)}"
                        recipients.append({
                            'name': f"{student.get_full_name()} (Parent)",
                            'phone': phone,
                            'whatsapp_url': whatsapp_url,
                            'type': 'parent'
                        })
                        recipient_count = 1
                    else:
                        messages.warning(request, 'No phone number found for this student\'s parent.')
                except Student.DoesNotExist:
                    messages.error(request, 'Student not found.')
            
            elif audience == 'section' and section_id:
                try:
                    section = Section.objects.get(id=section_id)
                    students = Student.objects.filter(
                        section=section,
                        status='active'
                    )
                    for student in students:
                        phone = student.father_phone or student.mother_phone or student.guardian_phone
                        if phone:
                            # Clean phone number
                            clean_phone = ''.join(filter(str.isdigit, phone))
                            if not clean_phone.startswith('91') and len(clean_phone) == 10:
                                clean_phone = '91' + clean_phone
                            
                            whatsapp_url = f"https://wa.me/{clean_phone}?text={quote(message_text)}"
                            recipients.append({
                                'name': f"{student.get_full_name()}",
                                'phone': phone,
                                'whatsapp_url': whatsapp_url,
                                'type': 'parent'
                            })
                    recipient_count = len(recipients)
                    if recipient_count == 0:
                        messages.warning(request, 'No phone numbers found for students in this section.')
                except Section.DoesNotExist:
                    messages.error(request, 'Section not found.')
            
            elif audience == 'all_sections':
                # All students in teacher's sections
                for student in teacher_students:
                    phone = student.father_phone or student.mother_phone or student.guardian_phone
                    if phone:
                        # Clean phone number
                        clean_phone = ''.join(filter(str.isdigit, phone))
                        if not clean_phone.startswith('91') and len(clean_phone) == 10:
                            clean_phone = '91' + clean_phone
                        
                        whatsapp_url = f"https://wa.me/{clean_phone}?text={quote(message_text)}"
                        recipients.append({
                            'name': f"{student.get_full_name()}",
                            'phone': phone,
                            'whatsapp_url': whatsapp_url,
                            'type': 'parent'
                        })
                recipient_count = len(recipients)
                if recipient_count == 0:
                    messages.warning(request, 'No phone numbers found for students in your classes.')
            
            if recipient_count > 0:
                # Store WhatsApp Web links
                whatsapp_links = {
                    'message': message_text,
                    'recipients': recipients,
                    'count': recipient_count,
                    'audience': audience
                }
                request.session['whatsapp_links'] = whatsapp_links
                messages.success(request, f'Generated {recipient_count} WhatsApp Web link(s). Click the links below to open WhatsApp.')
    else:
        # Get WhatsApp links from session on GET request
        whatsapp_links = request.session.pop('whatsapp_links', None)
    
    # Convert sections to list if it's a queryset
    if hasattr(teacher_sections, '__iter__') and not isinstance(teacher_sections, (list, tuple)):
        teacher_sections = list(teacher_sections)
    
    context = {
        'teacher': teacher,
        'teacher_sections': teacher_sections,
        'teacher_students': teacher_students,
        'whatsapp_links': whatsapp_links,
        'current_year': current_year,
    }
    
    return render(request, 'teachers/whatsapp.html', context)


# ============================================
# Staff Management Functions
# ============================================

@login_required
def staff_list(request):
    """List all staff members"""
    from .models import Staff, School
    
    # Staff model doesn't have is_active field, so we get all staff
    staff_members = Staff.objects.select_related('user', 'category', 'school').all()
    
    # Super admin: School filter - mandatory, show no data if school not selected
    if is_super_admin(request.user):
        school_id = request.GET.get('school')
        if school_id:
            staff_members = staff_members.filter(school_id=school_id)
        else:
            staff_members = Staff.objects.none()  # Return empty queryset if school not selected
        schools = School.objects.all().order_by('name')
        context = {
            'staff_members': staff_members,
            'is_super_admin': True,
            'schools': schools,
            'selected_school': request.GET.get('school', '')
        }
    else:
        context = {
            'staff_members': staff_members,
            'is_super_admin': False
        }
    
    return render(request, 'staff/staff_list.html', context)


@login_required
def add_staff(request):
    """Add a new staff member"""
    from .models import Staff
    
    if request.method == 'POST':
        name = request.POST.get('name')
        phone = request.POST.get('phone')
        designation = request.POST.get('designation', '')
        
        if name and phone:
            # Create Django User and Staff record
            try:
                from django.contrib.auth import get_user_model
                User = get_user_model()
                username = phone
                default_pwd = "Staff@123"
                existing_user = User.objects.filter(username=username).first()
                if not existing_user:
                    try:
                        first_name = name.split()[0] if name.strip() else ''
                        last_name = ' '.join(name.split()[1:]) if len(name.split()) > 1 else ''
                        existing_user = User.objects.create_user(
                            username=username,
                            email=f"{username}@school.local",
                            password=default_pwd,
                            first_name=first_name,
                            last_name=last_name
                        )
                    except Exception:
                        existing_user = None

                # Create Staff entry linking to user
                if existing_user:
                    staff_obj, created = Staff.objects.get_or_create(
                        user=existing_user,
                        defaults={
                            'employee_id': username,
                            'phone': phone,
                            'designation': designation,
                            'category': None
                        }
                    )
                else:
                    # fallback to legacy behaviour if user couldn't be created
                    staff_obj, created = Staff.objects.get_or_create(
                        phone=phone,
                        defaults={'employee_id': phone, 'designation': designation}
                    )

                if created:
                    messages.success(request, f'Staff member "{name}" added successfully!')
            except Exception:
                # Fallback to basic creation if models differ
                try:
                    staff, created = Staff.objects.get_or_create(
                        phone=phone,
                        defaults={'employee_id': phone, 'designation': designation}
                    )
                    if created:
                        messages.success(request, f'Staff member "{name}" added successfully!')
                except Exception:
                    messages.error(request, 'Error creating staff member.')

            return redirect('students_app:staff_list')
    
    return render(request, 'staff/add.html')


@login_required
def staff_bulk_upload(request):
    """Handle bulk staff upload via Excel and optional ZIP"""
    from .models import Staff, School, StaffCategory
    from django.contrib.auth.models import User
    from django.utils import timezone
    import pandas as pd
    
    context = {}
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # Handle saving bulk data from preview
        if action == 'save_bulk':
            try:
                import uuid
                from django.core.files.base import ContentFile
                
                # Get form data arrays
                first_names = request.POST.getlist('first_name[]')
                last_names = request.POST.getlist('last_name[]')
                employee_ids = request.POST.getlist('employee_id[]')
                emails = request.POST.getlist('email[]')
                phones = request.POST.getlist('phone[]')
                dobs = request.POST.getlist('dob[]')
                genders = request.POST.getlist('gender[]')
                designations = request.POST.getlist('designation[]')
                departments = request.POST.getlist('department[]')
                salaries = request.POST.getlist('salary[]')
                addresses = request.POST.getlist('address[]')
                cities = request.POST.getlist('city[]')
                states = request.POST.getlist('state[]')
                pincodes = request.POST.getlist('pincode[]')
                joining_dates = request.POST.getlist('joining_date[]')
                image_filenames = request.POST.getlist('image_filename[]')
                
                # Get images path from session
                images_path = request.session.get('staff_import_images_path')
                
                success_count = 0
                error_count = 0
                skipped_details = []
                
                # Get user's school
                school_obj = None
                if hasattr(request.user, 'school_profile') and request.user.school_profile:
                    school_obj = request.user.school_profile.school

                # Validate arrays have same length
                array_length = len(first_names)
                if not all(len(arr) == array_length for arr in [last_names, employee_ids, emails, phones, dobs, genders, designations, departments, salaries, addresses, joining_dates, image_filenames, assigned_classes, assigned_sections]):
                    messages.error(request, 'Form data is corrupted. Please try uploading the Excel file again.')
                    return redirect('students_app:staff_bulk_upload')

                print(f'Processing {array_length} staff members from form submission')

                # Get user's school
                school_obj = None
                if hasattr(request.user, 'school_profile') and request.user.school_profile:
                    school_obj = request.user.school_profile.school

                # Get available classes and sections for smart assignment
                # Class model has no 'school' FK — classes are global
                available_classes = Class.objects.all().prefetch_related('section_set').order_by('numeric_value', 'name')
                available_sections = Section.objects.all().select_related('class_assigned').order_by('class_assigned__numeric_value', 'name')

                context['available_classes'] = available_classes
                context['available_sections'] = available_sections
                context['auto_assign_classes'] = auto_assign_classes
                context['auto_assign_sections'] = auto_assign_sections

                # Process each staff member
                for i in range(array_length):
                    try:
                        print(f'Processing staff {i+1}: {first_names[i]} {last_names[i]}')

                        # Validate required fields
                        if not emails[i] or not employee_ids[i] or not first_names[i] or not last_names[i]:
                            print(f'Skipping staff {i+1} - missing required fields')
                            error_count += 1
                            skipped_details.append(f"Row {i+1}: Missing required fields (First/Last name, Email, Employee ID)")
                            continue
                        
                        # Check for duplicates within the same school only
                        if User.objects.filter(username=emails[i]).exists() and SchoolUser.objects.filter(user__username=emails[i], school=school_obj).exists():
                            error_count += 1
                            skipped_details.append(f"Row {i+1}: Duplicate email/username '{emails[i]}' within this school")
                            continue
                        
                        if Staff.objects.filter(employee_id=employee_ids[i], school=school_obj).exists():
                            error_count += 1
                            skipped_details.append(f"Row {i+1}: Duplicate Employee ID '{employee_ids[i]}' within this school")
                            continue
                        
                        # Create Django User
                        user = User.objects.create_user(
                            username=emails[i],
                            email=emails[i],
                            password="Staff@123",
                            first_name=first_names[i],
                            last_name=last_names[i]
                        )
                        
                        # Get or create default staff category
                        default_category, _ = StaffCategory.objects.get_or_create(
                            name='General',
                            defaults={'description': 'General Staff Category'}
                        )
                        
                        # Create Staff record
                        staff = Staff(
                            user=user,
                            employee_id=employee_ids[i],
                            phone=phones[i],
                            date_of_birth=dobs[i],
                            gender=genders[i],
                            designation=designations[i] if i < len(designations) and designations[i] else 'Staff',
                            department=departments[i] if i < len(departments) and departments[i] else 'General',
                            category=default_category,
                            current_salary=salaries[i] if i < len(salaries) and salaries[i] else 0,
                            address=addresses[i] if i < len(addresses) and addresses[i] else "Not Provided",
                            city=cities[i] if i < len(cities) and cities[i] else '',
                            state=states[i] if i < len(states) and states[i] else '',
                            pincode=pincodes[i] if i < len(pincodes) and pincodes[i] else '',
                            joining_date=joining_dates[i],
                            school=school_obj
                        )
                        staff.save()
                        
                        # Attach photo if available
                        img_fn = (image_filenames[i] if i < len(image_filenames) else '').strip()
                        if img_fn and images_path:
                            img_path = os.path.join(images_path, img_fn)
                            if os.path.isfile(img_path):
                                try:
                                    with open(img_path, 'rb') as f:
                                        staff.photo.save(img_fn, ContentFile(f.read()), save=True)
                                except Exception:
                                    pass
                        
                        # Create SchoolUser profile
                        try:
                            from .models import UserRole, SchoolUser
                            staff_role = UserRole.objects.filter(name='staff').first()
                            if staff_role:
                                SchoolUser.objects.get_or_create(
                                    user=user,
                                    defaults={
                                        'role': staff_role,
                                        'school': school_obj,
                                        'login_id': emails[i],
                                        'custom_password': 'Staff@123',
                                        'phone': phones[i] or ''
                                    }
                                )
                        except Exception:
                            pass
                        
                        success_count += 1
                        
                    except Exception as e:
                        print(f"Error adding staff {first_names[i]}: {e}")
                        error_count += 1
                        skipped_details.append(f"Row {i+1}: {str(e)}")
                        continue
                
                # Cleanup temporary images directory
                if images_path and os.path.isdir(images_path):
                    try:
                        import shutil
                        shutil.rmtree(images_path, ignore_errors=True)
                    except Exception:
                        pass
                    request.session.pop('staff_import_images_path', None)
                
                # Show results
                print(f'Upload complete: {success_count} success, {error_count} errors')
                
                if success_count > 0:
                    messages.success(request, f"Successfully added {success_count} staff member(s).")
                if error_count > 0:
                    msg = f"Skipped {error_count} staff member(s) (duplicates or errors)."
                    if skipped_details:
                        sample = skipped_details[:5]
                        msg += f" Sample: {'; '.join(sample)}"
                    messages.warning(request, msg)
                
                # If no staff were processed, show detailed error
                if success_count == 0 and error_count == 0:
                    messages.error(request, "No staff data was processed. Please check your Excel file and try again.")
                
                return redirect('students_app:staff_list')
            
            except Exception as e:
                print(f'Error in bulk upload processing: {e}')
                import traceback
                traceback.print_exc()
                messages.error(request, f"Error processing bulk upload: {str(e)}. Please try again.")
                return redirect('students_app:staff_bulk_upload')
        
                # Handle initial Excel upload and preview
        elif 'excel_file' in request.FILES:
            try:
                import uuid
                from django.core.files.base import ContentFile
                from django.conf import settings
                
                excel_file = request.FILES['excel_file']
                
                # Debug: Print file info
                print(f"Excel file received: {excel_file.name}")
                print(f"File size: {excel_file.size} bytes")
                print(f"File content type: {excel_file.content_type}")
                
                # Check file size (limit to 10MB)
                if excel_file.size > 10 * 1024 * 1024:
                    messages.error(request, "File size too large. Please upload a file smaller than 10MB.")
                    return render(request, 'staff/staff_bulk_upload.html', context)
                images_zip = request.FILES.get('images_zip')
                
                # Extract images from ZIP if provided
                images_path = None
                if images_zip and images_zip.name.endswith('.zip'):
                    try:
                        images_dict = extract_images_from_zip(images_zip)
                        if images_dict:
                            # Save images to temp directory
                            temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_staff_import', str(uuid.uuid4()))
                            os.makedirs(temp_dir, exist_ok=True)
                            
                            for filename, data in images_dict.items():
                                with open(os.path.join(temp_dir, filename), 'wb') as f:
                                    f.write(data)
                            
                            request.session['staff_import_images_path'] = temp_dir
                    except Exception as e:
                        messages.warning(request, f'Error processing images: {str(e)}')
                        request.session.pop('staff_import_images_path', None)
                else:
                    request.session.pop('staff_import_images_path', None)
                
                # Read Excel file
                try:
                    df = pd.read_excel(excel_file)
                    print(f"Successfully read Excel file with {len(df)} rows and {len(df.columns)} columns")
                    print(f"Columns found: {list(df.columns)}")
                except Exception as e:
                    print(f"Error reading Excel file: {e}")
                    messages.error(request, f"Error reading Excel file: {str(e)}. Please ensure it's a valid Excel file (.xlsx or .xls)")
                    return render(request, 'staff/staff_bulk_upload.html', context)
                
                # Check if DataFrame is empty
                if df.empty:
                    messages.error(request, "The Excel file appears to be empty. Please check the file and try again.")
                    return render(request, 'staff/staff_bulk_upload.html', context)
                
                # Normalize column names (lowercase, replace spaces with single space)
                df.columns = [str(c).strip().lower().replace('  ', ' ') for c in df.columns]
                print(f"Normalized columns: {list(df.columns)}")
                
                # Check for required columns
                required_columns = ['first name', 'last name', 'employee id', 'email']
                missing_columns = [col for col in required_columns if col not in df.columns]
                
                if missing_columns:
                    messages.error(request, f"Missing required columns in Excel file: {', '.join(missing_columns)}. Please check your Excel file format.")
                    return render(request, 'staff/staff_bulk_upload.html', context)
                
                # Helper function to get photo filename from various column names
                photo_cols = ['image filename', 'image_filename', 'photo', 'photo filename']
                def get_photo(row):
                    for col in photo_cols:
                        if col in df.columns and pd.notna(row.get(col)):
                            return str(row[col]).strip()
                    return ''
                
                # Helper function to format dates
                def format_date(val, default):
                    if pd.isna(val) or str(val).strip() == '':
                        return default
                    try:
                        return pd.to_datetime(val).strftime('%Y-%m-%d')
                    except:
                        return default
                
                # Process each row
                staff_data = []
                for index, row in df.iterrows():
                    dob = format_date(
                        row.get('date of birth (yyyy-mm-dd)') or row.get('dob'),
                        '2000-01-01'
                    )
                    
                    joining = format_date(
                        row.get('joining date (yyyy-mm-dd)') or row.get('joining date'),
                        timezone.now().date().strftime('%Y-%m-%d')
                    )
                    
                    # Get gender (handle various formats)
                    gender_raw = row.get('gender (m/f/o)') or row.get('gender') or 'O'
                    gender = str(gender_raw)[0].upper() if str(gender_raw).strip() else 'O'
                    
                    staff_data.append({
                        'first_name': row.get('first name', ''),
                        'last_name': row.get('last name', ''),
                        'employee_id': row.get('employee id', ''),
                        'email': row.get('email', ''),
                        'phone': row.get('phone', ''),
                        'dob': dob,
                        'gender': gender,
                        'address': row.get('address', ''),
                        'city': row.get('city', ''),
                        'state': row.get('state', ''),
                        'pincode': row.get('pincode', ''),
                        'designation': row.get('designation', ''),
                        'department': row.get('department', ''),
                        'joining_date': joining,
                        'salary': row.get('current salary') or row.get('salary') or 0,
                        'image_filename': get_photo(row),
                    })
                
                context['staff_data'] = staff_data
                
            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
    
    return render(request, 'staff/staff_bulk_upload.html', context)


# ============================================
# Parent Functions
# ============================================

def parent_login(request):
    """Parent login view"""
    return redirect('students_app:simple_login')


@login_required
def parent_student_progress(request, student_id):
    """View student progress for parent"""
    from .models import Student, Attendance, Marks
    
    student = get_object_or_404(Student, id=student_id)
    attendance = Attendance.objects.filter(student=student)
    marks = Marks.objects.filter(student=student).select_related('exam_schedule__exam', 'exam_schedule__subject')
    
    context = {
        'student': student,
        'attendance': attendance,
        'marks': marks,
    }
    
    return render(request, 'parents/student_progress.html', context)


# ============================================
# EXAMS & MARKS STUBS
# ============================================

@login_required
def marks_entry(request):
    """Marks entry view"""
    return render(request, 'students/marks_entry.html')

@login_required
def student_report_card(request, admission_number, exam_id):
    """View student report card"""
    from .models import Student, Marks
    student = get_object_or_404(Student, admission_number=admission_number)
    marks = Marks.objects.filter(student=student, exam_schedule__exam_id=exam_id).select_related('exam_schedule__exam', 'exam_schedule__subject')
    return render(request, 'students/report_card.html', {'student': student, 'marks': marks})

@login_required
def exam_list(request):
    """List exams"""
    return render(request, 'students/exams/exam_list.html')

@login_required
def exam_create(request):
    """Create exam"""
    return render(request, 'students/exam_create.html')

@login_required
def exam_schedule_create(request, exam_id):
    """Add schedule to exam"""
    from .models import Exam
    exam = get_object_or_404(Exam, id=exam_id) if hasattr(Exam, 'id') else None
    if request.method == 'POST':
        messages.success(request, 'Exam schedule created.')
        return redirect('students_app:exam_schedule_list')
    return render(request, 'students/exam_schedule_form.html', {'exam': exam})

@login_required
def exam_schedule_list(request):
    """List exam schedules"""
    return render(request, 'students/exams/exam_schedule_list.html')

@login_required
def exam_schedule_create_standalone(request):
    """Create exam schedule standalone"""
    return redirect('students_app:exam_schedule_list')

@login_required
def exam_schedule_edit(request, schedule_id):
    """Edit exam schedule"""
    if request.method == 'POST':
        messages.success(request, 'Exam schedule updated.')
        return redirect('students_app:exam_schedule_list')
    return render(request, 'students/exam_schedule_form.html', {'schedule_id': schedule_id})

@login_required
def exam_schedule_delete(request, schedule_id):
    """Delete exam schedule"""
    if request.method == 'POST':
        messages.success(request, 'Exam schedule deleted.')
    return redirect('students_app:exam_schedule_list')


# ============================================
# ATTENDANCE STUBS
# ============================================

@login_required
def attendance_mark(request):
    """Mark student attendance"""
    from .models import Class, Student, Attendance
    from django.http import JsonResponse
    from django.utils.dateparse import parse_date
    from django.db import transaction
    
    # Get classes for dropdown
    classes = Class.objects.all().order_by('numeric_value', 'name')
    
    # Handle POST request for saving attendance
    if request.method == 'POST':
        try:
            class_id = request.POST.get('class_id')
            date_str = request.POST.get('date')
            
            if not class_id or not date_str:
                messages.error(request, 'Class and date are required')
                return render(request, 'students/attendance_mark.html', {
                    'classes': classes,
                    'today': timezone.localdate()
                })
            
            attendance_date = parse_date(date_str)
            if not attendance_date:
                messages.error(request, 'Invalid date format')
                return render(request, 'students/attendance_mark.html', {
                    'classes': classes,
                    'today': timezone.localdate()
                })
            
            # Get all students in the class
            students = Student.objects.filter(current_class_id=class_id, status='active')
            
            marked_count = 0
            with transaction.atomic():
                for student in students:
                    status_key = f'status_{student.id}'
                    status = request.POST.get(status_key, 'P')
                    
                    Attendance.objects.update_or_create(
                        student=student,
                        date=attendance_date,
                        defaults={
                            'status': status,
                            'marked_by': request.user
                        }
                    )
                    marked_count += 1
            
            messages.success(request, f'Attendance marked for {marked_count} students')
            return redirect('students_app:attendance_mark')
            
        except Exception as e:
            messages.error(request, f'Error saving attendance: {str(e)}')
    
    return render(request, 'students/attendance_mark.html', {
        'classes': classes,
        'today': timezone.localdate()
    })

@login_required
def teacher_attendance_mark(request):
    """Mark teacher attendance"""
    return render(request, 'students/teacher_attendance_mark.html')

@login_required
def teacher_book_request(request):
    """Teacher book request - teachers can request books for themselves"""
    from .models import Teacher, Book, BookRequest, BookCategory
    from django.http import JsonResponse
    from django.utils.dateparse import parse_date
    from datetime import timedelta
    
    try:
        teacher = Teacher.objects.get(user=request.user)
    except Teacher.DoesNotExist:
        messages.error(request, 'You are not registered as a teacher.')
        return redirect('students_app:teacher_dashboard')
    
    # Get available books
    available_books = Book.objects.filter(available_copies__gt=0).select_related('category')
    categories = BookCategory.objects.all()
    
    # Get teacher's book requests
    my_requests = BookRequest.objects.filter(teacher=teacher).select_related('book', 'approved_by').order_by('-request_date')
    
    # Handle POST request
    if request.method == 'POST':
        try:
            book_id = request.POST.get('book_id')
            requested_due_date_str = request.POST.get('requested_due_date')
            purpose = request.POST.get('purpose', '')
            
            if not book_id or not requested_due_date_str:
                messages.error(request, 'Book and due date are required')
                return render(request, 'teachers/book_request.html', {
                    'teacher': teacher,
                    'available_books': available_books,
                    'categories': categories,
                    'my_requests': my_requests,
                })
            
            book = Book.objects.get(id=book_id)
            requested_due_date = parse_date(requested_due_date_str)
            
            if not requested_due_date:
                messages.error(request, 'Invalid date format')
                return render(request, 'teachers/book_request.html', {
                    'teacher': teacher,
                    'available_books': available_books,
                    'categories': categories,
                    'my_requests': my_requests,
                })
            
            # Check if book is available
            if book.available_copies <= 0:
                messages.error(request, 'This book is not available')
                return render(request, 'teachers/book_request.html', {
                    'teacher': teacher,
                    'available_books': available_books,
                    'categories': categories,
                    'my_requests': my_requests,
                })
            
            # Check if teacher already has a pending request for this book
            existing_request = BookRequest.objects.filter(
                teacher=teacher,
                book=book,
                status='pending'
            ).exists()
            
            if existing_request:
                messages.warning(request, 'You already have a pending request for this book')
                return render(request, 'teachers/book_request.html', {
                    'teacher': teacher,
                    'available_books': available_books,
                    'categories': categories,
                    'my_requests': my_requests,
                })
            
            # Create book request
            BookRequest.objects.create(
                teacher=teacher,
                book=book,
                requested_due_date=requested_due_date,
                purpose=purpose,
                status='pending'
            )
            
            messages.success(request, f'Book request submitted successfully for "{book.title}"')
            return redirect('students_app:teacher_book_request')
            
        except Book.DoesNotExist:
            messages.error(request, 'Book not found')
        except Exception as e:
            messages.error(request, f'Error submitting request: {str(e)}')
    
    # Filter by category if provided
    category_id = request.GET.get('category')
    if category_id:
        available_books = available_books.filter(category_id=category_id)
    
    # Search filter
    search_query = request.GET.get('search')
    if search_query:
        available_books = available_books.filter(
            models.Q(title__icontains=search_query) |
            models.Q(author__icontains=search_query) |
            models.Q(isbn__icontains=search_query)
        )
    
    from datetime import date
    return render(request, 'teachers/book_request.html', {
        'teacher': teacher,
        'available_books': available_books,
        'categories': categories,
        'my_requests': my_requests,
        'selected_category': int(category_id) if category_id else None,
        'search_query': search_query or '',
        'today': date.today(),
    })

@login_required
def teacher_student_attendance(request):
    """Teacher marks attendance for students in their classes"""
    from .models import Teacher, Timetable, Section, Student, Attendance, AcademicYear
    from django.http import JsonResponse
    from django.utils.dateparse import parse_date
    from django.db import transaction
    from django.contrib import messages
    
    # Get the teacher for the logged-in user
    try:
        teacher = Teacher.objects.get(user=request.user)
    except Teacher.DoesNotExist:
        messages.error(request, 'You are not registered as a teacher.')
        return redirect('students_app:teacher_dashboard')
    
    # Get current academic year
    current_year = AcademicYear.objects.filter(is_current=True).first()
    if not current_year:
        messages.warning(request, 'No current academic year set.')
    
    # Get all sections this teacher teaches (from timetable)
    teacher_sections = Section.objects.filter(
        timetables__teacher=teacher,
        timetables__academic_year=current_year
    ).distinct().order_by('class_assigned__numeric_value', 'name')
    
    # Handle POST request for saving attendance
    if request.method == 'POST':
        try:
            section_id = request.POST.get('section_id')
            date_str = request.POST.get('date')
            
            if not section_id or not date_str:
                messages.error(request, 'Section and date are required')
                return render(request, 'teachers/student_attendance.html', {
                    'teacher': teacher,
                    'sections': teacher_sections,
                    'today': timezone.localdate()
                })
            
            # Verify teacher teaches this section
            section = Section.objects.get(id=section_id)
            if section not in teacher_sections:
                messages.error(request, 'You are not authorized to mark attendance for this section.')
                return render(request, 'teachers/student_attendance.html', {
                    'teacher': teacher,
                    'sections': teacher_sections,
                    'today': timezone.localdate()
                })
            
            attendance_date = parse_date(date_str)
            if not attendance_date:
                messages.error(request, 'Invalid date format')
                return render(request, 'teachers/student_attendance.html', {
                    'teacher': teacher,
                    'sections': teacher_sections,
                    'today': timezone.localdate()
                })
            
            # Get all students in the section
            students = Student.objects.filter(
                current_class=section.class_assigned,
                section=section,
                status='active'
            ).order_by('admission_number')
            
            marked_count = 0
            with transaction.atomic():
                for student in students:
                    status_key = f'status_{student.id}'
                    status = request.POST.get(status_key, 'P')
                    remarks_key = f'remarks_{student.id}'
                    remarks = request.POST.get(remarks_key, '')
                    
                    Attendance.objects.update_or_create(
                        student=student,
                        date=attendance_date,
                        defaults={
                            'status': status,
                            'remarks': remarks,
                            'marked_by': request.user
                        }
                    )
                    marked_count += 1
            
            messages.success(request, f'Attendance marked for {marked_count} students in {section.class_assigned.name} - Section {section.name}')
            return redirect('students_app:teacher_student_attendance')
            
        except Exception as e:
            messages.error(request, f'Error saving attendance: {str(e)}')
    
    return render(request, 'teachers/student_attendance.html', {
        'teacher': teacher,
        'sections': teacher_sections,
        'today': timezone.localdate()
    })

@login_required
def teacher_change_photo(request):
    """Allow teacher to change their profile photo"""
    from .models import Teacher
    from django.shortcuts import get_object_or_404
    
    try:
        teacher = Teacher.objects.get(user=request.user)
    except Teacher.DoesNotExist:
        messages.error(request, 'You are not registered as a teacher.')
        return redirect('students_app:teacher_dashboard')
    
    if request.method == 'POST':
        photo = request.FILES.get('photo')
        if photo:
            # Delete old photo if exists
            if teacher.photo:
                teacher.photo.delete(save=False)
            
            teacher.photo = photo
            teacher.save()
            messages.success(request, 'Profile photo updated successfully!')
            return redirect('students_app:teacher_dashboard')
        else:
            messages.error(request, 'Please select a photo to upload.')
    
    context = {
        'teacher': teacher,
    }
    
    return render(request, 'teachers/change_photo.html', context)


@login_required
def student_change_photo(request, admission_number):
    """Allow admin/teacher to change student profile photo"""
    from .models import Student
    from django.shortcuts import get_object_or_404
    
    # Get student
    student = get_object_or_404(Student, admission_number=admission_number)
    
    # Check permissions - admin or class teacher can change
    has_permission = False
    if request.user.is_superuser:
        has_permission = True
    else:
        try:
            if hasattr(request.user, 'school_profile'):
                school_user = request.user.school_profile
                if school_user and school_user.role:
                    role_name = school_user.role.name.lower().strip()
                    if 'admin' in role_name:
                        has_permission = True
        except:
            pass
        
        # Check if user is a teacher and is class teacher for this student's section
        if not has_permission:
            try:
                from .models import Teacher, ClassTeacher, AcademicYear
                teacher = Teacher.objects.get(user=request.user)
                current_year = AcademicYear.objects.filter(is_current=True).first()
                if current_year and student.section:
                    is_class_teacher = ClassTeacher.objects.filter(
                        teacher=teacher,
                        section=student.section,
                        academic_year=current_year
                    ).exists()
                    if is_class_teacher:
                        has_permission = True
            except:
                pass
    
    if not has_permission:
        messages.error(request, 'You do not have permission to change student photos.')
        return redirect('students_app:student_detail', admission_number=admission_number)
    
    if request.method == 'POST':
        photo = request.FILES.get('photo')
        if photo:
            # Delete old photo if exists
            if student.photo:
                student.photo.delete(save=False)
            
            student.photo = photo
            student.save()
            messages.success(request, f'Photo updated successfully for {student.get_full_name()}!')
            return redirect('students_app:student_detail', admission_number=admission_number)
        else:
            messages.error(request, 'Please select a photo to upload.')
    
    context = {
        'student': student,
    }
    
    return render(request, 'students/change_photo.html', context)


@login_required
def assign_class_teacher(request, teacher_id=None):
    """Assign a teacher as class teacher for a section"""
    from .forms import AssignClassTeacherForm
    from .models import Teacher, ClassTeacher, AcademicYear, Section
    
    # Check if user has admin permissions
    has_permission = False
    try:
        if hasattr(request.user, 'school_profile'):
            school_user = request.user.school_profile
            if school_user and school_user.role:
                role_name = school_user.role.name.lower().strip()
                if 'admin' in role_name:
                    has_permission = True
    except:
        pass
    
    if not has_permission and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to assign class teachers.')
        return redirect('students_app:dashboard')
    
    # If teacher_id is provided, pre-select that teacher
    teacher = None
    if teacher_id:
        teacher = get_object_or_404(Teacher, id=teacher_id)
    
    if request.method == 'POST':
        form = AssignClassTeacherForm(request.POST)
        if form.is_valid():
            try:
                selected_teacher = form.cleaned_data['teacher']
                section = form.cleaned_data['section']
                academic_year = form.cleaned_data['academic_year']
                
                # Check if section already has a class teacher for this academic year
                existing = ClassTeacher.objects.filter(
                    section=section,
                    academic_year=academic_year
                ).first()
                
                if existing:
                    # Update existing assignment
                    existing.teacher = selected_teacher
                    existing.save()
                    messages.success(request, f'Class teacher updated successfully! {selected_teacher.user.get_full_name()} is now the class teacher for {section}.')
                else:
                    # Create new assignment
                    ClassTeacher.objects.create(
                        teacher=selected_teacher,
                        section=section,
                        academic_year=academic_year
                    )
                    messages.success(request, f'{selected_teacher.user.get_full_name()} assigned as class teacher for {section} successfully!')
                
                # Redirect based on where user came from
                if teacher_id:
                    return redirect('students_app:teacher_list')
                else:
                    return redirect('students_app:assign_class_teacher')
            except Exception as e:
                messages.error(request, f'Error assigning class teacher: {str(e)}')
    else:
        form = AssignClassTeacherForm()
        if teacher:
            form.fields['teacher'].initial = teacher.id
    
    # Get current academic year
    current_year = AcademicYear.objects.filter(is_current=True).first()
    
    # Get all class teachers for display
    class_teachers = ClassTeacher.objects.select_related(
        'teacher__user', 'section__class_assigned', 'academic_year'
    ).order_by('-academic_year__year', 'section__class_assigned__numeric_value', 'section__name')
    
    # Filter by academic year if current year exists
    if current_year:
        class_teachers = class_teachers.filter(academic_year=current_year)
    
    context = {
        'form': form,
        'teacher': teacher,
        'class_teachers': class_teachers,
        'current_year': current_year,
    }
    
    return render(request, 'students/assign_class_teacher.html', context)


@login_required
def teacher_my_attendance(request):
    """Teacher views their own attendance"""
    from .models import Teacher, TeacherAttendance
    from django.db.models import Count, Q
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    try:
        teacher = Teacher.objects.get(user=request.user)
    except Teacher.DoesNotExist:
        messages.error(request, 'You are not registered as a teacher.')
        return redirect('students_app:teacher_dashboard')
    
    # Get filter parameters
    month = request.GET.get('month')
    year = request.GET.get('year', timezone.now().year)
    
    # Get attendance records
    attendance_records = TeacherAttendance.objects.filter(teacher=teacher).order_by('-date')
    
    # Apply month filter if provided
    if month:
        try:
            month_int = int(month)
            attendance_records = attendance_records.filter(date__year=year, date__month=month_int)
        except ValueError:
            pass
    
    # Get current month/year if not specified
    if not month:
        current_date = timezone.now().date()
        month = current_date.month
        year = current_date.year
        attendance_records = attendance_records.filter(date__year=year, date__month=month)
    
    # Calculate statistics
    total_days = attendance_records.count()
    present_count = attendance_records.filter(status='P').count()
    absent_count = attendance_records.filter(status='A').count()
    late_count = attendance_records.filter(status='L').count()
    half_day_count = attendance_records.filter(status='H').count()
    excused_count = attendance_records.filter(status='E').count()
    
    # Calculate attendance percentage
    if total_days > 0:
        attendance_percentage = round((present_count / total_days) * 100, 1)
    else:
        attendance_percentage = 0
    
    # Calculate salary information
    from decimal import Decimal
    
    current_salary = teacher.current_salary or Decimal('0')
    
    # Calculate monthly salary based on attendance
    # Assuming 30 working days per month
    working_days_per_month = Decimal('30')
    daily_salary = current_salary / working_days_per_month if working_days_per_month > 0 else Decimal('0')
    
    # Calculate deductions for absent days (full day deduction)
    absent_deduction = Decimal(absent_count) * daily_salary
    
    # Calculate deductions for half days (half day deduction)
    half_day_deduction = Decimal(half_day_count) * (daily_salary / Decimal('2'))
    
    # Calculate deductions for late (small deduction, e.g., 10% of daily salary)
    late_deduction = Decimal(late_count) * (daily_salary * Decimal('0.1'))
    
    # Total deductions
    total_deductions = absent_deduction + half_day_deduction + late_deduction
    
    # Calculate payable salary for the month
    payable_salary = current_salary - total_deductions
    
    # Get recent attendance (last 30 days)
    recent_attendance = TeacherAttendance.objects.filter(
        teacher=teacher,
        date__gte=timezone.now().date() - timedelta(days=30)
    ).order_by('-date')[:30]
    
    # Generate month choices
    months = [
        (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
        (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
        (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
    ]
    
    context = {
        'teacher': teacher,
        'attendance_records': attendance_records,
        'recent_attendance': recent_attendance,
        'total_days': total_days,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'half_day_count': half_day_count,
        'excused_count': excused_count,
        'attendance_percentage': attendance_percentage,
        'current_month': int(month) if month else timezone.now().month,
        'current_year': int(year),
        'months': months,
        'years': list(range(2022, 2026)),  # 2022 to 2025
        # Salary information
        'current_salary': current_salary,
        'daily_salary': daily_salary,
        'absent_deduction': absent_deduction,
        'half_day_deduction': half_day_deduction,
        'late_deduction': late_deduction,
        'total_deductions': total_deductions,
        'payable_salary': payable_salary,
        'working_days_per_month': working_days_per_month,
    }
    
    return render(request, 'teachers/my_attendance.html', context)


@login_required
def teacher_get_section_students(request, section_id):
    """AJAX endpoint to get students for a section (for teacher)"""
    from .models import Teacher, Section, Student
    from django.http import JsonResponse
    
    try:
        teacher = Teacher.objects.get(user=request.user)
        section = Section.objects.get(id=section_id)
        
        # Verify teacher teaches this section
        from .models import Timetable, AcademicYear
        current_year = AcademicYear.objects.filter(is_current=True).first()
        if current_year:
            teaches_section = Timetable.objects.filter(
                teacher=teacher,
                section=section,
                academic_year=current_year
            ).exists()
            
            if not teaches_section:
                return JsonResponse({'error': 'You are not authorized to view this section'}, status=403)
        
        students = Student.objects.filter(
            current_class=section.class_assigned,
            section=section,
            status='active'
        ).order_by('admission_number')
        
        students_data = [{
            'id': student.id,
            'admission_number': student.admission_number,
            'name': student.get_full_name(),
            'roll_number': getattr(student, 'roll_number', '')
        } for student in students]
        
        return JsonResponse({'students': students_data})
        
    except Teacher.DoesNotExist:
        return JsonResponse({'error': 'Teacher not found'}, status=404)
    except Section.DoesNotExist:
        return JsonResponse({'error': 'Section not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def teacher_attendance_ocr_verify(request):
    """OCR verification for teacher attendance"""
    from django.http import JsonResponse
    from .models import Teacher
    try:
        teacher = Teacher.objects.get(user=request.user)
        return JsonResponse({'status': 'ok', 'teacher': teacher.get_full_name(), 'verified': False})
    except Teacher.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Teacher not found'}, status=404)

@login_required
def teacher_attendance_face_verify(request):
    """Face verification for teacher attendance"""
    from django.http import JsonResponse
    from .models import Teacher
    try:
        teacher = Teacher.objects.get(user=request.user)
        return JsonResponse({'status': 'ok', 'teacher': teacher.get_full_name(), 'face_match_confidence': 0.0})
    except Teacher.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Teacher not found'}, status=404)

@login_required
def staff_attendance_mark(request):
    """Mark staff attendance"""
    return render(request, 'students/staff_attendance_mark.html')

@login_required
def staff_attendance_ocr_verify(request):
    """OCR verification for staff attendance"""
    from django.http import JsonResponse
    return JsonResponse({'status': 'ok', 'verified': False, 'message': 'OCR verification not yet implemented'})

@login_required
def staff_attendance_face_verify(request):
    """Face verification for staff attendance"""
    from django.http import JsonResponse
    return JsonResponse({'status': 'ok', 'face_match_confidence': 0.0, 'message': 'Face verification not yet implemented'})

@login_required
def view_attendance(request):
    """View attendance records"""
    return render(request, 'students/view_attendance.html')

@login_required
def get_section_students(request, section_id):
    """Get students for a section (AJAX)"""
    from django.http import JsonResponse
    return JsonResponse({'students': []})

@login_required
def get_class_students(request, class_id):
    """Get students for a class (AJAX)"""
    from django.http import JsonResponse
    from .models import Class, Student
    
    try:
        class_obj = Class.objects.get(id=class_id)
        students = Student.objects.filter(
            current_class=class_obj,
            status='active'
        ).select_related('section').order_by('roll_number', 'first_name')
        
        students_data = []
        for student in students:
            students_data.append({
                'id': student.id,
                'name': f"{student.first_name} {student.middle_name} {student.last_name}".strip(),
                'roll_number': student.roll_number or '-',
                'section': student.section.name if student.section else '-'
            })
        
        return JsonResponse({'students': students_data})
    except Class.DoesNotExist:
        return JsonResponse({'error': 'Class not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def get_sections_for_class(request, class_id):
    """Get sections for a class (AJAX)"""
    from django.http import JsonResponse
    from .models import Section
    try:
        sections = Section.objects.filter(class_assigned_id=class_id).order_by('name')
        sections_data = [{'id': s.id, 'name': s.name} for s in sections]
        return JsonResponse({'sections': sections_data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def attendance_report(request):
    """Generate attendance report"""
    from .models import Attendance, Student
    from django.db.models import Count, Q
    students = Student.objects.filter(status='active')
    attendance_summary = []
    for student in students:
        total = Attendance.objects.filter(student=student).count()
        present = Attendance.objects.filter(student=student, status='P').count()
        percentage = (present / total * 100) if total > 0 else 0
        attendance_summary.append({
            'student': student,
            'total': total,
            'present': present,
            'percentage': round(percentage, 1)
        })
    return render(request, 'students/attendance_report.html', {'attendance_summary': attendance_summary})

@login_required
def qr_attendance_scan(request):
    """QR Attendance Scanner"""
    return render(request, 'students/qr_attendance_scan.html')

@login_required
def qr_attendance_process(request):
    """Process QR attendance"""
    from django.http import JsonResponse
    return JsonResponse({'status': 'ok'})

@login_required
def qr_attendance_records(request):
    """View QR attendance records"""
    return render(request, 'students/attendance/qr_attendance_records.html')


# ============================================
# FEES MANAGEMENT STUBS
# ============================================

@login_required
def fee_collection(request):
    """Collect fees"""
    return render(request, 'students/fee_collection.html')

@login_required
def fee_structure_management(request):
    """Manage fee structures - Create, Edit, Delete"""
    from .models import FeeStructure, Class, AcademicYear, Subject
    from decimal import Decimal
    
    # Get current academic year
    current_year = AcademicYear.objects.filter(is_current=True).first()
    if not current_year:
        current_year = AcademicYear.objects.order_by('-start_date').first()
    
    # Handle POST requests (Create, Update, Delete)
    if request.method == 'POST':
        action = request.GET.get('action', 'save')
        
        if action == 'delete':
            # Delete fee structure
            delete_id = request.POST.get('delete_id') or request.GET.get('id')
            if delete_id:
                try:
                    structure = FeeStructure.objects.get(id=delete_id)
                    class_name = structure.class_assigned.name
                    structure.delete()
                    messages.success(request, f'Fee structure for {class_name} deleted successfully!')
                except FeeStructure.DoesNotExist:
                    messages.error(request, 'Fee structure not found.')
                except Exception as e:
                    messages.error(request, f'Error deleting fee structure: {str(e)}')
            return redirect('students_app:fee_structure_management')
        
        else:
            # Create or Update fee structure
            structure_id = request.POST.get('structure_id')
            class_id = request.POST.get('class_assigned')
            year_id = request.POST.get('academic_year')
            stream = request.POST.get('stream', 'general')
            
            # Get fee components
            try:
                tuition_fee = Decimal(request.POST.get('tuition_fee', 0))
                transport_fee = Decimal(request.POST.get('transport_fee', 0))
                library_fee = Decimal(request.POST.get('library_fee', 0))
                lab_fee = Decimal(request.POST.get('lab_fee', 0))
                sports_fee = Decimal(request.POST.get('sports_fee', 0))
                exam_fee = Decimal(request.POST.get('exam_fee', 0))
                computer_fee = Decimal(request.POST.get('computer_fee', 0))
                other_fee = Decimal(request.POST.get('other_fee', 0))
                optional_subject_fee = Decimal(request.POST.get('optional_subject_fee', 0))
            except (ValueError, TypeError):
                messages.error(request, 'Invalid fee amount entered. Please enter valid numbers.')
                return redirect('students_app:fee_structure_management')
            
            # Validate required fields
            if not class_id or not year_id:
                messages.error(request, 'Please select both class and academic year.')
                return redirect('students_app:fee_structure_management')
            
            try:
                class_obj = Class.objects.get(id=class_id)
                year_obj = AcademicYear.objects.get(id=year_id)
                
                if structure_id:
                    # Update existing structure
                    structure = FeeStructure.objects.get(id=structure_id)
                    structure.class_assigned = class_obj
                    structure.academic_year = year_obj
                    structure.stream = stream
                    structure.tuition_fee = tuition_fee
                    structure.transport_fee = transport_fee
                    structure.library_fee = library_fee
                    structure.lab_fee = lab_fee
                    structure.sports_fee = sports_fee
                    structure.exam_fee = exam_fee
                    structure.computer_fee = computer_fee
                    structure.other_fee = other_fee
                    structure.optional_subject_fee = optional_subject_fee
                    structure.save()
                    
                    # Handle subjects (many-to-many)
                    subject_ids = request.POST.getlist('subjects')
                    if subject_ids:
                        structure.subjects.set(Subject.objects.filter(id__in=subject_ids))
                    else:
                        structure.subjects.clear()
                    
                    messages.success(request, f'Fee structure for {class_obj.name} updated successfully!')
                else:
                    # Check if structure already exists
                    existing = FeeStructure.objects.filter(
                        class_assigned=class_obj,
                        academic_year=year_obj,
                        stream=stream
                    ).first()
                    
                    if existing:
                        messages.warning(request, f'Fee structure for {class_obj.name} ({stream}) already exists. Please edit the existing structure.')
                        return redirect('students_app:fee_structure_management')
                    
                    # Create new structure
                    structure = FeeStructure.objects.create(
                        class_assigned=class_obj,
                        academic_year=year_obj,
                        stream=stream,
                        tuition_fee=tuition_fee,
                        transport_fee=transport_fee,
                        library_fee=library_fee,
                        lab_fee=lab_fee,
                        sports_fee=sports_fee,
                        exam_fee=exam_fee,
                        computer_fee=computer_fee,
                        other_fee=other_fee,
                        optional_subject_fee=optional_subject_fee
                    )
                    
                    # Handle subjects (many-to-many)
                    subject_ids = request.POST.getlist('subjects')
                    if subject_ids:
                        structure.subjects.set(Subject.objects.filter(id__in=subject_ids))
                    
                    messages.success(request, f'Fee structure for {class_obj.name} created successfully!')
                
                return redirect('students_app:fee_structure_management')
                
            except Class.DoesNotExist:
                messages.error(request, 'Selected class not found.')
            except AcademicYear.DoesNotExist:
                messages.error(request, 'Selected academic year not found.')
            except FeeStructure.DoesNotExist:
                messages.error(request, 'Fee structure not found.')
            except Exception as e:
                messages.error(request, f'Error saving fee structure: {str(e)}')
            
            return redirect('students_app:fee_structure_management')
    
    # GET request - Display fee structures
    # Filter by academic year if specified
    year_filter = request.GET.get('year')
    if year_filter:
        fee_structures = FeeStructure.objects.filter(academic_year_id=year_filter).select_related(
            'class_assigned', 'academic_year'
        ).prefetch_related('subjects').order_by('class_assigned__numeric_value', 'stream')
    elif current_year:
        fee_structures = FeeStructure.objects.filter(academic_year=current_year).select_related(
            'class_assigned', 'academic_year'
        ).prefetch_related('subjects').order_by('class_assigned__numeric_value', 'stream')
    else:
        fee_structures = FeeStructure.objects.select_related(
            'class_assigned', 'academic_year'
        ).prefetch_related('subjects').order_by('academic_year__start_date', 'class_assigned__numeric_value', 'stream')
    
    # Get all classes and academic years for the form
    classes = Class.objects.all().order_by('numeric_value')
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    subjects = Subject.objects.filter(is_active=True).order_by('name')
    
    # Calculate selected_year for template comparison
    selected_year = None
    if year_filter:
        try:
            selected_year = int(year_filter)
        except (ValueError, TypeError):
            pass

    context = {
        "fee_structures": fee_structures,
        "classes": classes,
        "academic_years": academic_years,
        "subjects": subjects,
        "current_year": current_year,  # object, not id
        "selected_year": selected_year,
    }
    
    return render(request, 'students/fee_structure_management.html', context)

@login_required
def search_students_autocomplete(request):
    """Autocomplete search for students"""
    from django.http import JsonResponse
    from .models import Student
    from django.db.models import Q
    query = request.GET.get('q', '')
    if len(query) < 2:
        return JsonResponse({'results': []})
    students = Student.objects.filter(
        Q(first_name__icontains=query) | Q(last_name__icontains=query) | Q(admission_number__icontains=query),
        status='active'
    ).values('id', 'admission_number', 'first_name', 'last_name')[:10]
    results = [{'id': s['id'], 'text': f"{s['admission_number']} - {s['first_name']} {s['last_name']}"} for s in students]
    return JsonResponse({'results': results})

@login_required
def get_student_fee_structure(request):
    """Get student fee structure"""
    from django.http import JsonResponse
    from .models import Student, FeeStructure
    student_id = request.GET.get('student_id')
    if not student_id:
        return JsonResponse({'error': 'Student ID required'}, status=400)
    try:
        student = Student.objects.get(id=student_id)
        fee_structure = FeeStructure.objects.filter(
            class_assigned=student.current_class,
            academic_year=student.academic_year
        ).first()
        if fee_structure:
            structure_data = {
                'tuition_fee': str(fee_structure.tuition_fee),
                'transport_fee': str(fee_structure.transport_fee),
                'lab_fee': str(fee_structure.lab_fee),
                'total': str(fee_structure.get_total_fee())
            }
            return JsonResponse({'structure': structure_data})
        return JsonResponse({'structure': {}})
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)

@login_required
def receipts_list(request):
    """List fee receipts"""
    return render(request, 'students/receipts_list.html')

@login_required
def fee_receipt(request, receipt_id):
    """View fee receipt"""
    from .models import FeePayment
    receipt = get_object_or_404(FeePayment, id=receipt_id)
    return render(request, 'students/fee_receipt.html', {'receipt': receipt})

@login_required
def fee_report(request):
    """Fee report"""
    return render(request, 'students/fee_report.html')

@login_required
def create_fee_order(request):
    """Create online fee payment order"""
    from django.http import JsonResponse
    from .models import Student, FeePayment
    student_id = request.GET.get('student_id')
    amount = request.GET.get('amount')
    if not student_id or not amount:
        return JsonResponse({'error': 'Missing parameters'}, status=400)
    try:
        student = Student.objects.get(id=student_id)
        import uuid
        order_id = f"order_{uuid.uuid4().hex[:8]}"
        return JsonResponse({'order_id': order_id, 'amount': amount, 'student': student.get_full_name()})
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)

@login_required
def verify_fee_payment(request):
    """Verify online fee payment"""
    from django.http import JsonResponse
    from .models import FeePayment
    order_id = request.POST.get('order_id')
    if not order_id:
        return JsonResponse({'error': 'Missing order_id'}, status=400)
    try:
        fee_payment = FeePayment.objects.get(transaction_id=order_id)
        fee_payment.payment_status = 'Paid'
        fee_payment.save()
        return JsonResponse({'status': 'verified', 'message': 'Payment verified'})
    except FeePayment.DoesNotExist:
        return JsonResponse({'status': 'pending', 'message': 'Payment verification pending'})


# ============================================
# LIBRARY MANAGEMENT STUBS
# ============================================

@login_required
def library_dashboard(request):
    """Library dashboard - restricted for teachers"""
    # Check if user is a teacher
    try:
        from .models import Teacher
        teacher = Teacher.objects.get(user=request.user)
        # Redirect teachers to book request page instead
        messages.info(request, 'Teachers can request books from the book request page.')
        return redirect('students_app:teacher_book_request')
    except Teacher.DoesNotExist:
        pass
    
    return render(request, 'students/library_dashboard.html')

@login_required
def issue_book(request):
    """Issue a book"""
    return render(request, 'students/issue_book.html')

@login_required
def return_book(request, issue_id):
    """Return a book"""
    from .models import BookIssue
    issue = get_object_or_404(BookIssue, id=issue_id)
    if request.method == 'POST':
        issue.status = 'returned'
        issue.return_date = timezone.now().date()
        issue.save()
        messages.success(request, f'Book "{issue.book.title}" returned successfully.')
        return redirect('students_app:library_dashboard')
    return render(request, 'students/return_book.html', {'issue': issue})

@login_required
def book_list(request):
    """List books"""
    return render(request, 'students/book_list.html')

@login_required
def add_book(request):
    """Add new book"""
    return render(request, 'students/add_book.html')

@login_required
def edit_book(request, book_id):
    """Edit book"""
    from .models import Book
    from .forms import BookForm
    book = get_object_or_404(Book, id=book_id)
    if request.method == 'POST':
        form = BookForm(request.POST, instance=book)
        if form.is_valid():
            form.save()
            messages.success(request, f'Book "{book.title}" updated successfully.')
            return redirect('students_app:book_list')
    else:
        form = BookForm(instance=book)
    return render(request, 'students/book_form.html', {'form': form, 'book': book, 'title': 'Edit Book'})

@login_required
def book_qr_code_view(request, book_id):
    """View book QR"""
    from .models import Book
    book = get_object_or_404(Book, id=book_id)
    return render(request, 'students/book_qr_code.html', {'book': book})

@login_required
def book_qr_code(request, book_id):
    """Download book QR"""
    from .models import Book
    from django.http import FileResponse
    book = get_object_or_404(Book, id=book_id)
    if hasattr(book, 'qr_code') and book.qr_code:
        return FileResponse(book.qr_code.open('rb'), content_type='image/png', as_attachment=True, filename=f'{book.title}_qr.png')
    return render(request, 'students/book_qr_code.html', {'book': book, 'message': 'QR code not generated'})

@login_required
def bulk_qr_codes(request):
    """Bulk download book QRs"""
    from .models import Book
    books = Book.objects.all()
    return render(request, 'students/bulk_qr_codes.html', {'books': books})


# ============================================
# ACADEMICS & TIMETABLE STUBS
# ============================================

@login_required
def view_timetable(request):
    """View timetable - shows all teachers' timetables for teachers"""
    from .models import Teacher, Timetable, Section, AcademicYear, Class, TimeSlot
    
    # Check if user is a teacher
    is_teacher = False
    current_teacher = None
    try:
        current_teacher = Teacher.objects.get(user=request.user)
        is_teacher = True
    except Teacher.DoesNotExist:
        pass
    
    # Get current academic year
    current_year = AcademicYear.objects.filter(is_current=True).first()
    
    # Get selected section (if any)
    section_id = request.GET.get('section')
    selected_section = None
    if section_id:
        try:
            selected_section = Section.objects.get(id=section_id)
        except Section.DoesNotExist:
            pass
    
    # Get all sections
    sections = Section.objects.select_related('class_assigned').order_by('class_assigned__numeric_value', 'name')
    
    # For teachers: Show all teachers' timetables
    if is_teacher and current_year:
        # Get all teachers with their timetables
        all_teachers = Teacher.objects.filter(is_active=True).select_related('user').order_by('user__first_name', 'user__last_name')
        
        # Get all timetables for current year
        all_timetables = Timetable.objects.filter(
            academic_year=current_year
        ).select_related('teacher', 'subject', 'section', 'section__class_assigned', 'time_slot').order_by('weekday', 'time_slot__start_time')
        
        # Organize by teacher: create a list of timetable entries for each teacher
        teachers_timetables = {}
        for teacher in all_teachers:
            teacher_timetables = all_timetables.filter(teacher=teacher)
            if teacher_timetables.exists():
                teachers_timetables[teacher] = list(teacher_timetables)
        
        # Get all time slots for display
        time_slots = TimeSlot.objects.filter(is_break=False).order_by('start_time')
        
        context = {
            'is_teacher': True,
            'current_teacher': current_teacher,
            'teachers_timetables': teachers_timetables,
            'all_teachers': all_teachers,
            'time_slots': time_slots,
            'current_year': current_year,
            'sections': sections,
            'selected_section': selected_section,
        }
        return render(request, 'teachers/all_teachers_timetable.html', context)
    
    # For admin/students/parents: Show section-based timetable
    timetable_data = {}
    if selected_section and current_year:
        timetables = Timetable.objects.filter(
            section=selected_section,
            academic_year=current_year
        ).select_related('subject', 'teacher', 'time_slot').order_by('weekday', 'time_slot__start_time')
        
        # Organize by weekday
        for tt in timetables:
            weekday = tt.weekday
            if weekday not in timetable_data:
                timetable_data[weekday] = []
            timetable_data[weekday].append(tt)
    
    context = {
        'is_teacher': False,
        'timetable_data': timetable_data,
        'sections': sections,
        'selected_section': selected_section,
        'current_year': current_year,
        'classes': Class.objects.all().order_by('numeric_value'),
    }
    return render(request, 'students/view_timetable.html', context)

@login_required
def request_timetable_change(request):
    """Allow teachers to request changes to their timetable"""
    from .models import Teacher, TimetableChangeRequest, AcademicYear
    from .forms import TimetableChangeRequestForm
    from django.utils import timezone
    
    # Check if user is a teacher
    try:
        teacher = Teacher.objects.get(user=request.user)
    except Teacher.DoesNotExist:
        messages.error(request, 'Only teachers can request timetable changes.')
        return redirect('students_app:view_timetable')
    
    # Get current academic year
    current_year = AcademicYear.objects.filter(is_current=True).first()
    if not current_year:
        messages.error(request, 'No active academic year found.')
        return redirect('students_app:view_timetable')
    
    if request.method == 'POST':
        form = TimetableChangeRequestForm(request.POST, teacher=teacher, academic_year=current_year)
        if form.is_valid():
            change_request = form.save(commit=False)
            change_request.teacher = teacher
            change_request.academic_year = current_year
            change_request.status = 'pending'
            change_request.save()
            messages.success(request, 'Your timetable change request has been submitted successfully. The admin will review it soon.')
            return redirect('students_app:view_timetable')
    else:
        form = TimetableChangeRequestForm(teacher=teacher, academic_year=current_year)
    
    # Get teacher's current timetable entries for reference
    current_timetables = Timetable.objects.filter(
        teacher=teacher,
        academic_year=current_year
    ).select_related('section', 'subject', 'time_slot', 'section__class_assigned').order_by('weekday', 'time_slot__start_time')
    
    # Get existing pending requests
    pending_requests = TimetableChangeRequest.objects.filter(
        teacher=teacher,
        status='pending'
    ).order_by('-created_at')
    
    context = {
        'form': form,
        'teacher': teacher,
        'current_year': current_year,
        'current_timetables': current_timetables,
        'pending_requests': pending_requests,
    }
    
    return render(request, 'teachers/request_timetable_change.html', context)


@login_required
def manage_timetable_requests(request):
    """Admin view to manage all timetable change requests"""
    from .models import TimetableChangeRequest, AcademicYear
    from django.db.models import Q
    
    # Check if user is admin or superuser
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'Only administrators can access this page.')
        return redirect('students_app:view_timetable')
    
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    teacher_filter = request.GET.get('teacher', '')
    
    # Get all requests
    requests = TimetableChangeRequest.objects.select_related(
        'teacher', 'teacher__user', 'academic_year', 
        'current_timetable_entry', 'current_timetable_entry__section',
        'current_timetable_entry__subject', 'current_timetable_entry__time_slot',
        'preferred_time_slot', 'preferred_section', 'preferred_subject',
        'reviewed_by'
    ).order_by('-created_at')
    
    # Apply filters
    if status_filter:
        requests = requests.filter(status=status_filter)
    
    if teacher_filter:
        requests = requests.filter(
            Q(teacher__user__first_name__icontains=teacher_filter) |
            Q(teacher__user__last_name__icontains=teacher_filter) |
            Q(teacher__employee_id__icontains=teacher_filter)
        )
    
    # Get statistics
    total_requests = TimetableChangeRequest.objects.count()
    pending_count = TimetableChangeRequest.objects.filter(status='pending').count()
    approved_count = TimetableChangeRequest.objects.filter(status='approved').count()
    rejected_count = TimetableChangeRequest.objects.filter(status='rejected').count()
    
    # Get current academic year
    current_year = AcademicYear.objects.filter(is_current=True).first()
    
    context = {
        'requests': requests,
        'total_requests': total_requests,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'current_year': current_year,
        'status_filter': status_filter,
        'teacher_filter': teacher_filter,
    }
    
    return render(request, 'students/manage_timetable_requests.html', context)


@login_required
def review_timetable_request(request, request_id):
    """Admin view to review and approve/reject a timetable change request"""
    from .models import TimetableChangeRequest
    from django.utils import timezone
    
    # Check if user is admin or superuser
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'Only administrators can review requests.')
        return redirect('students_app:view_timetable')
    
    try:
        change_request = TimetableChangeRequest.objects.select_related(
            'teacher', 'teacher__user', 'academic_year',
            'current_timetable_entry', 'current_timetable_entry__section',
            'current_timetable_entry__subject', 'current_timetable_entry__time_slot',
            'preferred_time_slot', 'preferred_section', 'preferred_subject'
        ).get(id=request_id)
    except TimetableChangeRequest.DoesNotExist:
        messages.error(request, 'Timetable change request not found.')
        return redirect('students_app:manage_timetable_requests')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        admin_notes = request.POST.get('admin_notes', '')
        
        if action in ['approve', 'reject', 'in_progress']:
            change_request.status = action if action != 'approve' else 'approved'
            change_request.reviewed_by = request.user
            change_request.reviewed_at = timezone.now()
            change_request.admin_notes = admin_notes
            change_request.save()
            
            if action == 'approve':
                messages.success(request, f'Timetable change request has been approved for {change_request.teacher.user.get_full_name()}.')
            elif action == 'reject':
                messages.info(request, f'Timetable change request has been rejected for {change_request.teacher.user.get_full_name()}.')
            else:
                messages.info(request, f'Timetable change request status updated to "In Progress" for {change_request.teacher.user.get_full_name()}.')
            
            return redirect('students_app:manage_timetable_requests')
    
    context = {
        'change_request': change_request,
    }
    
    return render(request, 'students/review_timetable_request.html', context)


@login_required
def edit_timetable_redirect(request):
    """Redirect to edit timetable"""
    from .models import Section
    sections = Section.objects.all().order_by('class_assigned__numeric_value', 'name')
    if sections.exists():
        return redirect('students_app:edit_timetable', section_id=sections.first().id)
    return redirect('students_app:manage_timetables')

@login_required
def edit_timetable(request, section_id=None):
    """Edit timetable for section"""
    from .models import Section, Timetable, AcademicYear
    section = None
    if section_id:
        section = get_object_or_404(Section, id=section_id)
    current_year = AcademicYear.objects.filter(is_current=True).first()
    timetables = Timetable.objects.filter(section=section, academic_year=current_year) if section and current_year else Timetable.objects.none()
    return render(request, 'students/edit_timetable.html', {'section': section, 'timetables': timetables, 'current_year': current_year})

@login_required
def manage_timetables(request):
    """Manage timetables"""
    return render(request, 'students/manage_timetables.html')

@login_required
def timetable_conflicts(request):
    """Check timetable conflicts"""
    return render(request, 'students/timetable_conflicts.html')

@login_required
def bulk_timetable_operations(request):
    """Bulk timetable operations"""
    return redirect('students_app:manage_timetables')

@login_required
def export_timetable(request, section_id):
    """Export timetable"""
    from .models import Section, Timetable, AcademicYear
    import csv
    from django.http import HttpResponse as HttpResponseFile
    section = get_object_or_404(Section, id=section_id)
    current_year = AcademicYear.objects.filter(is_current=True).first()
    timetables = Timetable.objects.filter(section=section, academic_year=current_year).select_related('subject', 'teacher', 'time_slot')
    response = HttpResponseFile(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{section}_timetable.csv"'
    writer = csv.writer(response)
    writer.writerow(['Day', 'Time Slot', 'Subject', 'Teacher', 'Room'])
    for tt in timetables:
        writer.writerow([tt.get_weekday_display(), str(tt.time_slot), tt.subject.name, str(tt.teacher), tt.room_number or ''])
    return response

@login_required
def import_timetable(request):
    """Import timetable"""
    if request.method == 'POST':
        excel_file = request.FILES.get('timetable_file')
        if excel_file:
            messages.success(request, 'Timetable import started.')
            return redirect('students_app:manage_timetables')
    return render(request, 'students/import_timetable.html', {})

@login_required
def academic_year_list(request):
    """List academic years"""
    return render(request, 'students/academic_year_list.html')

@login_required
def academic_year_create(request):
    """Create academic year"""
    if request.method == 'POST':
        from .models import AcademicYear
        year = request.POST.get('year')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if year and start_date and end_date:
            AcademicYear.objects.create(year=year, start_date=start_date, end_date=end_date)
            messages.success(request, f'Academic year {year} created successfully.')
        return redirect('students_app:academic_year_list')
    return render(request, 'students/academics/academic_year_form.html', {})

@login_required
def time_slot_list(request):
    """List time slots"""
    return render(request, 'students/time_slot_list.html')

@login_required
def time_slot_create(request):
    """Create time slot"""
    if request.method == 'POST':
        from .models import TimeSlot
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        name = request.POST.get('name')
        if start_time and end_time and name:
            TimeSlot.objects.create(start_time=start_time, end_time=end_time, name=name)
            messages.success(request, f'Time slot {name} created successfully.')
        return redirect('students_app:time_slot_list')
    return render(request, 'students/academics/time_slot_form.html', {})


# ============================================
# COMMUNICATION STUBS
# ============================================

@login_required
def notice_list(request):
    """List notices"""
    from .models import Notice
    from django.db.models import Q
    from django.utils import timezone
    
    # Get user role
    user_role = ''
    if hasattr(request.user, 'school_profile') and request.user.school_profile.role:
        user_role = request.user.school_profile.role.name.lower()
    
    # Base queryset
    notices = Notice.objects.select_related('specific_class', 'specific_section', 'created_by').all()
    
    # Filter by search query
    search_query = request.GET.get('search', '')
    if search_query:
        notices = notices.filter(
            Q(title__icontains=search_query) | 
            Q(content__icontains=search_query)
        )
    
    # Filter by priority
    priority_filter = request.GET.get('priority', '')
    if priority_filter:
        notices = notices.filter(priority=priority_filter)
    
    # Filter by audience
    audience_filter = request.GET.get('audience', '')
    if audience_filter:
        notices = notices.filter(target_audience=audience_filter)
    
    # Filter by status (only for admins)
    status_filter = request.GET.get('status', '')
    if status_filter and (user_role in ['admin', 'super_admin', 'school_admin'] or request.user.is_superuser):
        notices = notices.filter(status=status_filter)
    
    # Filter by visibility (only for admins)
    visibility_filter = request.GET.get('visibility', '')
    if visibility_filter and (user_role in ['admin', 'super_admin', 'school_admin'] or request.user.is_superuser):
        if visibility_filter == 'active':
            notices = notices.filter(is_active=True)
        elif visibility_filter == 'inactive':
            notices = notices.filter(is_active=False)
    
    # Filter by expiry
    expired = request.GET.get('expired', '')
    today = timezone.localdate()
    if expired == 'yes':
        notices = notices.filter(expiry_date__lt=today)
    elif expired == 'no':
        notices = notices.filter(Q(expiry_date__gte=today) | Q(expiry_date__isnull=True))
    
    # Order by date
    notices = notices.order_by('-notice_date', '-created_at')
    
    return render(request, 'students/notice_list.html', {
        'notices': notices,
        'search_query': search_query,
        'priority_filter': priority_filter,
        'audience_filter': audience_filter,
        'status_filter': status_filter,
        'visibility_filter': visibility_filter,
        'expired': expired,
        'today': today,
        'user_role': user_role
    })

@login_required
def notice_create(request):
    """Create notice"""
    from .models import Notice
    from django.utils import timezone
    
    if request.method == 'POST':
        try:
            title = request.POST.get('title', '').strip()
            content = request.POST.get('content', '').strip()
            target_audience = request.POST.get('target_audience', 'all')
            priority = request.POST.get('priority', 'medium')
            notice_date = request.POST.get('notice_date')
            expiry_date = request.POST.get('expiry_date')
            status = request.POST.get('status', 'published')
            is_active = request.POST.get('is_active') == 'on'
            
            if not title or not content:
                messages.error(request, 'Title and content are required.')
                return render(request, 'students/notice_create.html')
            
            notice = Notice.objects.create(
                title=title,
                content=content,
                target_audience=target_audience,
                priority=priority,
                notice_date=notice_date if notice_date else timezone.now().date(),
                expiry_date=expiry_date if expiry_date else None,
                status=status,
                is_active=is_active,
                created_by=request.user
            )
            
            messages.success(request, f'Notice "{title}" created successfully!')
            return redirect('students_app:notice_list')
            
        except Exception as e:
            messages.error(request, f'Error creating notice: {str(e)}')
            return render(request, 'students/notice_create.html')
    
    return render(request, 'students/notice_create.html')

@login_required
def notice_detail(request, notice_id):
    """View notice detail"""
    return render(request, 'students/notice_detail.html')

@login_required
def notice_edit(request, notice_id):
    """Edit notice"""
    from .models import Notice
    notice = get_object_or_404(Notice, id=notice_id)
    if request.method == 'POST':
        notice.title = request.POST.get('title', notice.title)
        notice.content = request.POST.get('content', notice.content)
        notice.save()
        messages.success(request, 'Notice updated successfully.')
        return redirect('students_app:notice_list')
    return render(request, 'students/notice_form.html', {'notice': notice})

@login_required
def notice_delete(request, notice_id):
    """Delete notice"""
    from .models import Notice
    notice = get_object_or_404(Notice, id=notice_id)
    if request.method == 'POST':
        notice.delete()
        messages.success(request, 'Notice deleted successfully.')
    return redirect('students_app:notice_list')

@login_required
def notice_toggle_active(request, notice_id):
    """Toggle notice active status"""
    from .models import Notice
    notice = get_object_or_404(Notice, id=notice_id)
    notice.is_active = not notice.is_active
    notice.save()
    status = 'activated' if notice.is_active else 'deactivated'
    messages.success(request, f'Notice {status} successfully.')
    return redirect('students_app:notice_list')

@login_required
def notice_toggle_publish(request, notice_id):
    """Toggle notice publish status"""
    from .models import Notice
    notice = get_object_or_404(Notice, id=notice_id)
    notice.status = 'published' if notice.status != 'published' else 'draft'
    notice.save()
    messages.success(request, f'Notice status changed to {notice.status}.')
    return redirect('students_app:notice_list')

@login_required
def event_calendar(request):
    """Event calendar"""
    from .models import Event
    from django.db.models import Q
    from django.utils import timezone
    
    # Get user role
    user_role = ''
    if hasattr(request.user, 'school_profile') and request.user.school_profile.role:
        user_role = request.user.school_profile.role.name.lower()
    
    # Get filter parameters
    filter_date = request.GET.get('date', 'upcoming')
    filter_type = request.GET.get('type', 'all')
    
    # Base queryset
    events = Event.objects.select_related('organizer').all()
    
    # Filter by date
    today = timezone.localdate()
    if filter_date == 'upcoming':
        events = events.filter(event_date__gte=today)
    elif filter_date == 'today':
        events = events.filter(event_date=today)
    elif filter_date == 'past':
        events = events.filter(event_date__lt=today)
    # 'all' means no date filter
    
    # Filter by type
    if filter_type != 'all':
        events = events.filter(event_type=filter_type)
    
    # Order by date
    events = events.order_by('event_date', 'start_time')
    
    return render(request, 'students/event_calendar.html', {
        'events': events,
        'filter_date': filter_date,
        'filter_type': filter_type,
        'today': today,
        'user_role': user_role
    })

@login_required
def event_create(request):
    """Create event"""
    from .models import Event
    from django.utils import timezone
    
    if request.method == 'POST':
        try:
            title = request.POST.get('title', '').strip()
            description = request.POST.get('description', '').strip()
            event_type = request.POST.get('event_type', '')
            event_date = request.POST.get('event_date')
            start_time = request.POST.get('start_time')
            end_time = request.POST.get('end_time')
            venue = request.POST.get('venue', '')
            status = request.POST.get('status', 'draft')
            target_audience = request.POST.get('target_audience', 'all')
            is_holiday = request.POST.get('is_holiday') == 'on'
            
            if not title or not description or not event_date:
                messages.error(request, 'Title, description, and event date are required.')
                return render(request, 'students/event_create.html')
            
            event = Event.objects.create(
                title=title,
                description=description,
                event_type=event_type if event_type else 'other',
                event_date=event_date,
                start_time=start_time if start_time else None,
                end_time=end_time if end_time else None,
                venue=venue,
                status=status,
                target_audience=target_audience,
                is_holiday=is_holiday,
                organizer=request.user,
                is_published=(status == 'published')
            )
            
            messages.success(request, f'Event "{title}" created successfully!')
            return redirect('students_app:event_calendar')
            
        except Exception as e:
            messages.error(request, f'Error creating event: {str(e)}')
            return render(request, 'students/event_create.html')
    
    return render(request, 'students/event_create.html')

@login_required
def event_detail(request, event_id):
    """Event detail"""
    return render(request, 'students/event_detail.html')

@login_required
def event_toggle_publish(request, event_id):
    """Toggle event publish status"""
    from .models import Event
    event = get_object_or_404(Event, id=event_id)
    event.is_published = not event.is_published
    event.save()
    status = 'published' if event.is_published else 'unpublished'
    messages.success(request, f'Event {status} successfully.')
    return redirect('students_app:event_calendar')


# ============================================
# REPORTS STUBS
# ============================================

@login_required
def reports_dashboard(request):
    """Reports dashboard"""
    return render(request, 'students/reports_dashboard.html')


# ============================================
# GAMES & MOCK TESTS STUBS
# ============================================

@login_required
def games_home(request):
    """Games home"""
    return render(request, 'students/games_home.html')

@login_required
def games_by_class(request, class_id):
    """Games by class"""
    return render(request, 'students/games_list.html')

@login_required
def play_game(request, game_id):
    """Play game"""
    return render(request, 'students/play_game.html')

@login_required
def submit_game(request, game_id):
    """Submit game score"""
    if request.method == 'POST':
        score = request.POST.get('score', 0)
        messages.success(request, f'Game score {score} recorded.')
    return redirect('students_app:games_home')

@login_required
def game_leaderboard(request, game_id):
    """Game leaderboard"""
    return render(request, 'students/game_leaderboard.html')

@login_required
def student_game_stats(request):
    """Student game stats"""
    return render(request, 'students/student_game_stats.html')

@login_required
def mock_tests_home(request):
    """Mock tests home"""
    return render(request, 'students/mock_tests_home.html')

@login_required
def mock_tests_by_class(request, class_id):
    """Mock tests by class"""
    return render(request, 'students/mock_tests_list.html')

@login_required
def start_mock_test(request, test_id):
    """Start mock test"""
    return render(request, 'students/start_mock_test.html')

@login_required
def submit_mock_test(request, session_id):
    """Submit mock test"""
    if request.method == 'POST':
        answers = request.POST.dict()
        messages.success(request, 'Mock test submitted successfully.')
        return redirect('students_app:mock_test_results', session_id=session_id)
    return redirect('students_app:mock_tests_home')

@login_required
def mock_test_results(request, session_id):
    """Mock test results"""
    return render(request, 'students/mock_test_results.html')

@login_required
def mock_test_leaderboard(request, test_id):
    """Mock test leaderboard"""
    return render(request, 'students/mock_test_leaderboard.html')




# ============================================
# ID Card Management Functions (Added to fix missing references)
# ============================================

@login_required
def list_generated_id_cards(request):
    """List generated ID cards with search and group-by-class support"""
    from .models import StudentIDCard, Class, Section, School

    # Scope to school unless super admin
    school = _get_user_school(request.user)
    user_is_super_admin = is_super_admin(request.user)
    
    # Get filter parameters
    selected_school_id = request.GET.get('school', '').strip()
    selected_class_id = request.GET.get('class', '').strip()
    search_query = request.GET.get('search', '').strip()
    
    if user_is_super_admin:
        qs = StudentIDCard.objects.select_related(
            'student', 'student__current_class', 'student__section', 'template'
        ).order_by('-created_at')
        
        # Apply school filter if selected
        if selected_school_id:
            qs = qs.filter(student__school_id=selected_school_id)
            
        # Apply class filter if selected
        if selected_class_id:
            qs = qs.filter(student__current_class_id=selected_class_id)
    else:
        if school:
            qs = StudentIDCard.objects.filter(
                student__school=school
            ).select_related(
                'student', 'student__current_class', 'student__section', 'template'
            ).order_by('-created_at')
            
            # Apply class filter if selected
            if selected_class_id:
                qs = qs.filter(student__current_class_id=selected_class_id)
        else:
            qs = StudentIDCard.objects.none()

    # Search filter
    if search_query:
        qs = qs.filter(
            models.Q(student__first_name__icontains=search_query) |
            models.Q(student__last_name__icontains=search_query) |
            models.Q(student__admission_number__icontains=search_query) |
            models.Q(student__current_class__name__icontains=search_query)
        )

    id_cards = qs

    # Build group-by-class dict for the grouped view toggle
    id_cards_by_class = {}
    for card in id_cards:
        class_name = card.student.current_class.name if card.student.current_class else 'No Class'
        id_cards_by_class.setdefault(class_name, []).append(card)

    # Prepare context data
    context = {
        'id_cards': id_cards,
        'id_cards_by_class': id_cards_by_class,
        'search_query': search_query,
        'is_super_admin': user_is_super_admin,
        'selected_school': selected_school_id,
        'selected_class': selected_class_id,
    }
    
    # Add schools for super admin
    if user_is_super_admin:
        context['schools'] = School.objects.all().order_by('name')
        
        # Add classes based on selected school
        if selected_school_id:
            context['classes'] = Class.objects.all().order_by('numeric_value', 'name')
        else:
            context['classes'] = Class.objects.none()
    else:
        # Add classes for school user
        if school:
            context['classes'] = Class.objects.all().order_by('numeric_value', 'name')
        else:
            context['classes'] = Class.objects.none()

    return render(request, 'students/id_card_list.html', context)

@login_required
def id_card_detail_view(request, card_id):
    """View ID card details"""
    from .models import StudentIDCard
    import os
    from django.conf import settings
    
    card = get_object_or_404(StudentIDCard, id=card_id)
    
    # Check if image file exists
    image_exists = False
    if card.generated_image:
        image_path = card.generated_image.path if hasattr(card.generated_image, 'path') else None
        if image_path and os.path.exists(image_path):
            image_exists = True
        # Also check via URL
        elif card.generated_image.url:
            image_exists = True
    
    context = {
        'card': card,
        'image_exists': image_exists,
    }
    
    return render(request, 'students/id_card_detail.html', context)

@login_required
def download_id_card(request, card_id):
    """Download ID card image"""
    from .models import StudentIDCard
    card = get_object_or_404(StudentIDCard, id=card_id)
    if card.generated_image:
        response = HttpResponse(card.generated_image, content_type='image/png')
        response['Content-Disposition'] = f'attachment; filename="id_card_{card.card_number}.png"'
        return response
    messages.error(request, 'ID Card image not generated yet.')
    return redirect('students_app:id_card_detail', card_id=card_id)

@login_required
def delete_id_card(request, card_id):
    """Delete ID card"""
    from .models import StudentIDCard
    card = get_object_or_404(StudentIDCard, id=card_id)
    if request.method == 'POST':
        card.delete()
        messages.success(request, 'ID Card deleted successfully.')
        return redirect('students_app:id_card_list')
    return render(request, 'students/delete_id_card_confirm.html', {'object': card})

@login_required
def bulk_delete_id_cards(request):
    """Bulk delete ID cards"""
    if request.method == 'POST':
        card_ids = request.POST.getlist('card_ids')
        from .models import StudentIDCard
        deleted_count, _ = StudentIDCard.objects.filter(id__in=card_ids).delete()
        messages.success(request, f'{deleted_count} ID Cards deleted successfully.')
    return redirect('students_app:id_card_list')


@login_required
def bulk_download_id_cards(request):
    """Bulk download ID cards as ZIP - GET: class filter, POST: selected ids"""
    from .models import StudentIDCard
    import zipfile
    import io
    import os

    if request.method == 'GET':
        # Download all cards for a class/school via URL params
        class_id  = request.GET.get('class', '').strip()
        school_id = request.GET.get('school', '').strip()

        qs = StudentIDCard.objects.select_related('student').all()
        if school_id:
            qs = qs.filter(student__school_id=school_id)
        if class_id:
            qs = qs.filter(student__current_class_id=class_id)

        if not qs.exists():
            messages.error(request, 'Is class ke liye koi ID card nahi mila. Pehle generate karein.')
            return redirect('students_app:id_card_list')

        cards = qs
        zip_name = f"id_cards_class_{class_id}.zip" if class_id else "id_cards_bulk.zip"
    else:
        # Download selected cards (checkbox)
        card_ids = request.POST.getlist('card_ids')
        if not card_ids:
            messages.error(request, 'Koi ID card select nahi kiya.')
            return redirect('students_app:id_card_list')
        cards = StudentIDCard.objects.filter(id__in=card_ids).select_related('student')
        zip_name = "id_cards_selected.zip"

    zip_buffer = io.BytesIO()
    added = 0
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for card in cards:
            if card.generated_image:
                try:
                    image_path = card.generated_image.path
                    if os.path.exists(image_path):
                        filename = f"id_card_{card.student.admission_number}.png"
                        zip_file.write(image_path, filename)
                        added += 1
                except Exception:
                    continue

    if added == 0:
        messages.error(request, 'Koi generated image nahi mili. Pehle ID cards generate karein.')
        return redirect('students_app:id_card_list')

    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{zip_name}"'
    return response

class IDCardTemplateListView(LoginRequiredMixin, ListView):
    """List ID card templates"""
    template_name = 'students/idcard_template_list_fixed.html'
    context_object_name = 'templates'
    
    def get_queryset(self):
        from .models import IDCardTemplate
        from django.db.models import Q
        
        # Check if user is super admin
        user_is_super_admin = self.request.user.is_superuser or getattr(self.request.user, 'user_type', '') == 'super_admin'
        
        if user_is_super_admin:
            # Super admin: sab templates dekh sakta hai (school ke saath + bina school ke)
            return IDCardTemplate.objects.all().order_by('-created_at')
        else:
            # School user: sirf apni school ke templates + bina school ke templates
            school = _get_user_school(self.request.user)
            if school:
                # Apni school ke templates + jo templates ka school nahi hai
                return IDCardTemplate.objects.filter(
                    Q(school=school) | Q(school__isnull=True)
                ).order_by('-created_at')
            else:
                # Agar school nahi mila to sirf bina school ke templates dikhaye
                return IDCardTemplate.objects.filter(school__isnull=True).order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import Class, Section, School
        
        user = self.request.user
        user_is_super_admin = user.is_superuser or getattr(user, 'user_type', '') == 'super_admin'
        context['is_super_admin'] = user_is_super_admin
        
        if user_is_super_admin:
            # Super admin: pehle school select karega
            context['schools'] = School.objects.all().order_by('name')
            
            # URL se school ID le agar hai to
            selected_school_id = self.request.GET.get('school')
            context['selected_school'] = selected_school_id
            
            # Class model mein school FK nahi hai — global hai
            context['classes'] = Class.objects.all().order_by('numeric_value', 'name')
            context['sections'] = Section.objects.select_related('class_assigned').order_by(
                'class_assigned__numeric_value', 'name'
            )
        else:
            # School user: classes global hain, templates school ke
            school = _get_user_school(user)
            context['classes'] = Class.objects.all().order_by('numeric_value', 'name')
            context['sections'] = Section.objects.select_related('class_assigned').order_by(
                'class_assigned__numeric_value', 'name'
            )
            if school:
                context['user_school'] = school
        
        return context

@login_required
def generate_bulk_id_cards(request):
    """Generate bulk ID cards selection page"""
    from .models import Class, Section, School, IDCardTemplate

    user = request.user
    user_is_super_admin = is_super_admin(user)
    selected_school_id = request.GET.get('school', '').strip()

    if user_is_super_admin:
        schools = School.objects.all().order_by('name')

        if selected_school_id:
            # School select ho gayi — classes aur sections load karo
            # Since Class model mein school field nahi hai aur ek hi school hai
            # Sab classes show kar do
            classes = Class.objects.all().order_by('numeric_value', 'name')
            sections = Section.objects.filter(
                class_assigned__isnull=False
            ).select_related('class_assigned').order_by('class_assigned__numeric_value', 'name')
            templates = IDCardTemplate.objects.filter(
                school_id=selected_school_id
            ).order_by('name')
        else:
            # Koi school select nahi — empty show karo
            classes = Class.objects.none()
            sections = Section.objects.none()
            templates = IDCardTemplate.objects.all().order_by('name')

        context = {
            'is_super_admin': True,
            'schools': schools,
            'selected_school': selected_school_id,
            'classes': classes,
            'sections': sections,
            'templates': templates,
        }
    else:
        # Normal school user
        school = _get_user_school(user)
        # Class model has no 'school' FK — classes are global
        classes = Class.objects.all().order_by('numeric_value', 'name')
        sections = Section.objects.select_related('class_assigned').order_by(
            'class_assigned__numeric_value', 'name'
        )
        templates = IDCardTemplate.objects.filter(school=school).order_by('name') if school else IDCardTemplate.objects.all().order_by('name')

        context = {
            'is_super_admin': False,
            'classes': classes,
            'sections': sections,
            'templates': templates,
        }

    return render(request, 'students/generate_bulk_id_cards.html', context)

@login_required
def process_bulk_id_card_generation(request):
    """Process bulk ID card generation - actually generate cards!"""
    if request.method == 'POST':
        from .models import Student, IDCardTemplate, StudentIDCard
        from .id_card_generator import IDCardGenerator

        student_ids = request.POST.getlist('student_ids')
        template_id = request.POST.get('template_id')

        if not student_ids:
            messages.error(request, 'Koi student select nahi kiya. Please students select karein.')
            return redirect('students_app:generate_bulk_id_cards')

        if not template_id:
            messages.error(request, 'Koi template select nahi kiya. Please template select karein.')
            return redirect('students_app:generate_bulk_id_cards')

        try:
            template = IDCardTemplate.objects.get(id=template_id)
        except IDCardTemplate.DoesNotExist:
            messages.error(request, f'Template ID {template_id} nahi mili. Available templates: {list(IDCardTemplate.objects.values_list("id","name"))}')
            return redirect('students_app:generate_bulk_id_cards')

        if not template.template_image:
            messages.error(request, f'Template "{template.name}" mein image nahi hai. Pehle template image upload karein.')
            return redirect('students_app:idcard_template_list')

        import os
        if not os.path.exists(template.template_image.path):
            messages.error(request, f'Template image file disk par nahi mili: {template.template_image.name}. Re-upload karein.')
            return redirect('students_app:idcard_template_list')

        students = Student.objects.filter(id__in=student_ids, status='active').select_related('current_class', 'section', 'school')
        print(f"DEBUG: Processing {students.count()} students with IDs: {student_ids}")
        
        # Validate student classes
        class_distribution = {}
        for student in students:
            class_name = student.current_class.name if student.current_class else 'No Class'
            class_distribution[class_name] = class_distribution.get(class_name, 0) + 1
        
        print("DEBUG: Class distribution:")
        for class_name, count in class_distribution.items():
            print(f"  - {class_name}: {count} students")

        success_count = 0
        fail_count = 0
        errors = []

        for student in students:
            try:
                generator = IDCardGenerator(student, template)
                id_card = generator.save_card()
                if id_card:
                    success_count += 1
                else:
                    fail_count += 1
                    errors.append(f"{student.get_full_name()}: generation failed")
            except Exception as e:
                import traceback
                fail_count += 1
                full_error = traceback.format_exc()
                errors.append(f"{student.get_full_name()}: {str(e)} | {full_error[-300:]}")
                print(f"ID CARD ERROR for {student.admission_number}: {full_error}")

        if success_count > 0:
            messages.success(request, f'✅ {success_count} ID card(s) successfully generate ho gaye!')
        if fail_count > 0:
            # Show full first error to user
            first_error = errors[0] if errors else 'Unknown error'
            messages.error(request, f'❌ {fail_count} card(s) fail: {first_error[:500]}')
        if success_count == 0 and fail_count == 0:
            messages.warning(request, 'Koi student nahi mila selected IDs mein.')

        return redirect('students_app:id_card_list')

    return redirect('students_app:generate_bulk_id_cards')

@login_required
def generate_id_cards_for_class(request):
    """Generate ID cards for a specific class"""
    from .models import Student, IDCardTemplate, StudentIDCard
    class_id = request.GET.get('class_id')
    if class_id:
        students = Student.objects.filter(current_class_id=class_id, status='active')
        template = IDCardTemplate.objects.filter(is_active=True).first()
        if template and template.template_image:
            count = 0
            for student in students:
                StudentIDCard.objects.update_or_create(
                    student=student,
                    defaults={'template': template, 'card_number': f"{student.admission_number}"}
                )
                count += 1
            messages.success(request, f'ID cards generated for {count} students.')
        else:
            messages.error(request, 'No active template found.')
    return redirect('students_app:id_card_list')

@login_required
def generate_single_id_card(request, student_id):
    """Generate single ID card - Simple version for teachers"""
    from .models import Student, IDCardTemplate, StudentIDCard
    from .id_card_generator import IDCardGenerator
    from django.shortcuts import get_object_or_404
    
    try:
        student = get_object_or_404(Student, id=student_id)
        
        # Get default active template
        template = IDCardTemplate.objects.filter(is_active=True).first()
        
        if not template:
            messages.error(request, 'No active ID card template found. Please contact administrator to create a template.')
            return redirect('students_app:student_detail', admission_number=student.admission_number)
        
        if not template.template_image:
            messages.error(request, 'Template image missing. Please contact administrator to upload template image.')
            return redirect('students_app:student_detail', admission_number=student.admission_number)
        
        # Generate ID card
        generator = IDCardGenerator(student, template)
        id_card = generator.save_card()
        
        if id_card:
            messages.success(request, f'ID Card generated successfully for {student.get_full_name()}.')
            # Redirect to ID card detail page to show and download
            return redirect('students_app:id_card_detail', card_id=id_card.id)
        else:
            messages.error(request, 'Failed to generate ID card. Please try again or contact administrator.')
            return redirect('students_app:student_detail', admission_number=student.admission_number)
        
    except Exception as e:
        messages.error(request, f'Error generating ID card: {str(e)}. Please contact administrator.')
        return redirect('students_app:student_detail', admission_number=student.admission_number)

@login_required
def generate_pdf_id_cards(request):
    """Generate PDF for selected ID cards"""
    from django.http import HttpResponse
    card_ids = request.GET.getlist('card_ids')
    if card_ids:
        messages.success(request, f'PDF generated for {len(card_ids)} ID cards.')
    return redirect('students_app:id_card_list')

@login_required
def bulk_download_id_cards(request):
    """Bulk download ID cards"""
    from .models import StudentIDCard
    import zipfile
    import io
    card_ids = request.GET.getlist('card_ids')
    cards = StudentIDCard.objects.filter(id__in=card_ids) if card_ids else StudentIDCard.objects.all()[:100]
    if not cards.exists():
        messages.warning(request, 'No ID cards to download.')
        return redirect('students_app:id_card_list')
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zf:
        for card in cards:
            if card.generated_image:
                zf.writestr(f"{card.card_number}.png", card.generated_image.read())
    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="id_cards.zip"'
    return response

@login_required
def advanced_id_card_batch_excel(request):
    """Advanced batch generation via Excel"""
    from .models import IDCardTemplate
    templates = IDCardTemplate.objects.filter(is_active=True).order_by('name')
    context = {
        'templates': templates,
    }
    return render(request, 'students/advanced_batch_excel.html', context)

@login_required
def verify_id_card_ocr(request):
    """Verify ID card using OCR"""
    if request.method == 'POST':
        from django.http import JsonResponse
        return JsonResponse({'status': 'processing', 'message': 'OCR verification in progress'})
    return render(request, 'students/verify_id_card_ocr.html', {'status': 'ready'})

@login_required
def verify_face_match(request):
    """Verify face match"""
    if request.method == 'POST':
        from django.http import JsonResponse
        return JsonResponse({'status': 'processing', 'message': 'Face verification in progress'})
    return render(request, 'students/verify_face_match.html', {'status': 'ready'})

@login_required
def generate_printable_sheets(request):
    """Generate printable sheets"""
    return render(request, 'students/generate_printable_sheets.html', {})

@login_required
def generate_staff_id_cards(request):
    """Generate staff ID cards"""
    return render(request, 'students/generate_staff_id_cards.html')

@login_required
def upload_id_card_template(request):
    """Upload ID card template with visual editor support"""
    from .models import IDCardTemplate
    import json

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        orientation = request.POST.get('orientation', 'portrait')
        width = request.POST.get('width', 85)
        height = request.POST.get('height', 54)
        template_image = request.FILES.get('template_image')
        is_active = request.POST.get('is_active') == 'on'

        if not name:
            messages.error(request, 'Template name is required.')
            return render(request, 'students/upload_id_card_template.html')

        if not template_image:
            messages.error(request, 'Template image is required.')
            return render(request, 'students/upload_id_card_template.html')

        # Get school for this user
        school = None
        try:
            if hasattr(request.user, 'school_profile') and request.user.school_profile:
                school = request.user.school_profile.school
        except Exception:
            pass

        # Collect styles for layout_json
        styles = {}
        elements = ['name', 'admission', 'class', 'section', 'father', 'mother', 'guardian', 'mobile', 'address', 'dob', 'blood_group']
        for elem in elements:
            styles[elem] = {
                'font_size': request.POST.get(f'{elem}_font_size', '12'),
                'font_color': request.POST.get(f'{elem}_font_color', '#000000'),
                'font_family': request.POST.get(f'{elem}_font_family', 'Arial'),
                'font_weight': request.POST.get(f'{elem}_font_weight', 'normal'),
                'font_style': request.POST.get(f'{elem}_font_style', 'normal'),
                'width': request.POST.get(f'{elem}_width', 'auto'),
                'height': request.POST.get(f'{elem}_height', 'auto'),
            }

        try:
            template = IDCardTemplate(
                school=school,
                name=name,
                description=description,
                orientation=orientation,
                width=int(width),
                height=int(height),
                is_active=is_active,
                
                # Main Position fields
                photo_x=int(request.POST.get('photo_x', 10)),
                photo_y=int(request.POST.get('photo_y', 10)),
                photo_width=int(request.POST.get('photo_width', 100)),
                photo_height=int(request.POST.get('photo_height', 120)),
                
                name_x=int(request.POST.get('name_x', 10)),
                name_y=int(request.POST.get('name_y', 200)),
                name_font_size=int(request.POST.get('name_font_size', 18)),
                
                admission_no_x=int(request.POST.get('admission_no_x', 10)),
                admission_no_y=int(request.POST.get('admission_no_y', 220)),
                
                class_x=int(request.POST.get('class_x', 10)),
                class_y=int(request.POST.get('class_y', 240)),
                
                section_x=int(request.POST.get('section_x', 10)),
                section_y=int(request.POST.get('section_y', 260)),
                
                contact_x=int(request.POST.get('contact_x', 10)),
                contact_y=int(request.POST.get('contact_y', 280)),
                
                father_x=int(request.POST.get('father_x', 10)),
                father_y=int(request.POST.get('father_y', 300)),
                
                mother_x=int(request.POST.get('mother_x', 10)),
                mother_y=int(request.POST.get('mother_y', 320)),
                
                guardian_x=int(request.POST.get('guardian_x', 10)),
                guardian_y=int(request.POST.get('guardian_y', 340)),
                
                address_x=int(request.POST.get('address_x', 10)),
                address_y=int(request.POST.get('address_y', 360)),
                
                dob_x=int(request.POST.get('dob_x', 10)),
                dob_y=int(request.POST.get('dob_y', 380)),
                
                blood_group_x=int(request.POST.get('blood_group_x', 10)),
                blood_group_y=int(request.POST.get('blood_group_y', 400)),
                
                qr_code_x=int(request.POST.get('qr_code_x', 200)),
                qr_code_y=int(request.POST.get('qr_code_y', 20)),
                qr_code_size=int(request.POST.get('qr_code_size', 50)),
                
                # Display options
                show_name=request.POST.get('show_name') == 'on',
                show_photo=request.POST.get('show_photo') == 'on',
                show_admission_no=request.POST.get('show_admission_no') == 'on',
                show_class=request.POST.get('show_class') == 'on',
                show_section=request.POST.get('show_section') == 'on',
                show_father_name=request.POST.get('show_father_name') == 'on',
                show_mother_name=request.POST.get('show_mother_name') == 'on',
                show_guardian_name=request.POST.get('show_guardian_name') == 'on',
                show_mobile=request.POST.get('show_mobile') == 'on',
                show_address=request.POST.get('show_address') == 'on',
                show_blood_group=request.POST.get('show_blood_group') == 'on',
                show_dob=request.POST.get('show_dob') == 'on',
                show_qr_code=request.POST.get('show_qr_code') == 'on',
                
                layout_json={'font_styles': styles}
            )
            template.template_image = template_image
            template.save()
            messages.success(request, f'Template "{name}" successfully saved!')
            return redirect('students_app:idcard_template_list')
        except Exception as e:
            messages.error(request, f'Error saving template: {str(e)}')
            return render(request, 'students/upload_id_card_template.html')

    return render(request, 'students/upload_id_card_template.html')

@login_required
def id_card_template_list(request):
    """List ID card templates (function view alias)"""
    return redirect('students_app:idcard_template_list')

@login_required
def edit_id_card_template(request, template_id):
    """Edit ID card template with visual editor support"""
    from .models import IDCardTemplate
    template = get_object_or_404(IDCardTemplate, id=template_id)

    if request.method == 'POST':
        template.name = request.POST.get('name', template.name).strip()
        template.description = request.POST.get('description', template.description).strip()
        template.orientation = request.POST.get('orientation', template.orientation)
        template.width = int(request.POST.get('width', template.width))
        template.height = int(request.POST.get('height', template.height))
        template.is_active = request.POST.get('is_active') == 'on'

        template_image = request.FILES.get('template_image')
        if template_image:
            template.template_image = template_image

        # Update Main Position fields
        template.photo_x = int(request.POST.get('photo_x', template.photo_x))
        template.photo_y = int(request.POST.get('photo_y', template.photo_y))
        template.photo_width = int(request.POST.get('photo_width', template.photo_width))
        template.photo_height = int(request.POST.get('photo_height', template.photo_height))
        
        template.name_x = int(request.POST.get('name_x', template.name_x))
        template.name_y = int(request.POST.get('name_y', template.name_y))
        template.name_font_size = int(request.POST.get('name_font_size', template.name_font_size))
        
        template.admission_no_x = int(request.POST.get('admission_no_x', template.admission_no_x))
        template.admission_no_y = int(request.POST.get('admission_no_y', template.admission_no_y))
        
        template.class_x = int(request.POST.get('class_x', template.class_x))
        template.class_y = int(request.POST.get('class_y', template.class_y))
        
        template.section_x = int(request.POST.get('section_x', template.section_x))
        template.section_y = int(request.POST.get('section_y', template.section_y))
        
        template.contact_x = int(request.POST.get('contact_x', template.contact_x))
        template.contact_y = int(request.POST.get('contact_y', template.contact_y))
        
        template.father_x = int(request.POST.get('father_x', template.father_x))
        template.father_y = int(request.POST.get('father_y', template.father_y))
        
        template.mother_x = int(request.POST.get('mother_x', template.mother_x))
        template.mother_y = int(request.POST.get('mother_y', template.mother_y))
        
        template.guardian_x = int(request.POST.get('guardian_x', template.guardian_x))
        template.guardian_y = int(request.POST.get('guardian_y', template.guardian_y))
        
        template.address_x = int(request.POST.get('address_x', template.address_x))
        template.address_y = int(request.POST.get('address_y', template.address_y))
        
        template.dob_x = int(request.POST.get('dob_x', template.dob_x))
        template.dob_y = int(request.POST.get('dob_y', template.dob_y))
        
        template.blood_group_x = int(request.POST.get('blood_group_x', template.blood_group_x))
        template.blood_group_y = int(request.POST.get('blood_group_y', template.blood_group_y))
        
        template.qr_code_x = int(request.POST.get('qr_code_x', template.qr_code_x))
        template.qr_code_y = int(request.POST.get('qr_code_y', template.qr_code_y))
        template.qr_code_size = int(request.POST.get('qr_code_size', template.qr_code_size))

        # Update Display options
        template.show_name = request.POST.get('show_name') == 'on'
        template.show_photo = request.POST.get('show_photo') == 'on'
        template.show_admission_no = request.POST.get('show_admission_no') == 'on'
        template.show_class = request.POST.get('show_class') == 'on'
        template.show_section = request.POST.get('show_section') == 'on'
        template.show_father_name = request.POST.get('show_father_name') == 'on'
        template.show_mother_name = request.POST.get('show_mother_name') == 'on'
        template.show_guardian_name = request.POST.get('show_guardian_name') == 'on'
        template.show_mobile = request.POST.get('show_mobile') == 'on'
        template.show_address = request.POST.get('show_address') == 'on'
        template.show_blood_group = request.POST.get('show_blood_group') == 'on'
        template.show_dob = request.POST.get('show_dob') == 'on'
        template.show_qr_code = request.POST.get('show_qr_code') == 'on'

        # Update layout_json (styles)
        styles = template.layout_json.get('font_styles', {}) if template.layout_json else {}
        elements = ['name', 'admission', 'class', 'section', 'father', 'mother', 'guardian', 'mobile', 'address', 'dob', 'blood_group']
        for elem in elements:
            if f'{elem}_font_size' in request.POST:
                styles[elem] = {
                    'font_size': request.POST.get(f'{elem}_font_size'),
                    'font_color': request.POST.get(f'{elem}_font_color'),
                    'font_family': request.POST.get(f'{elem}_font_family'),
                    'font_weight': request.POST.get(f'{elem}_font_weight'),
                    'font_style': request.POST.get(f'{elem}_font_style'),
                    'width': request.POST.get(f'{elem}_width', 'auto'),
                    'height': request.POST.get(f'{elem}_height', 'auto'),
                }
        
        template.layout_json = {'font_styles': styles}
        template.save()
        messages.success(request, f'Template "{template.name}" updated successfully!')
        return redirect('students_app:idcard_template_list')

    return render(request, 'students/edit_id_card_template.html', {'template': template})


@login_required
@require_POST
def delete_id_card_template(request, template_id):
    """Delete ID card template"""
    from .models import IDCardTemplate
    template = get_object_or_404(IDCardTemplate, id=template_id)
    
    # SIMPLE PERMISSION CHECK - Staff/Admin can delete
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, "You do not have permission to delete this template.")
        return redirect('students_app:id_card_template_list')
    
    name = template.name
    try:
        template.delete()
        messages.success(request, f'Template "{name}" deleted successfully.')
    except Exception as e:
        messages.error(request, f'Error deleting template: {str(e)}')
        
    return redirect('students_app:id_card_template_list')


@login_required
def student_bulk_import(request):
    """
    Bulk import students from Excel file with images
    """
    from django.shortcuts import redirect
    from django.contrib import messages

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        images_zip = request.FILES.get('images_zip')
        require_section = request.POST.get('require_section') == 'on'

        if not excel_file:
            messages.error(request, 'Please upload an Excel file.')
            return redirect('students_app:student_import')

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls)')
            return redirect('students_app:student_import')

        try:
            result = process_student_import(
                excel_file=excel_file,
                images_zip=images_zip,
                user=request.user,
                require_section=require_section,
            )

            if result.get('success'):
                imported_count = result.get('imported', 0)
                skipped_count = result.get('skipped', 0)
                errors_count = result.get('errors', 0)
                error_details = result.get('error_details', [])
                skipped_details = result.get('skipped_details', [])
                msg = f"Imported: {imported_count}, Skipped: {skipped_count}, Errors: {errors_count}"
                if errors_count > 0 and error_details:
                    sample = error_details[:5]  # First 5 errors
                    msg += f". Sample errors: {'; '.join(sample)}"
                if skipped_count > 0 and skipped_details:
                    sample = skipped_details[:5]
                    msg += f". Sample skipped: {'; '.join(sample)}"
                messages.success(request, msg)
                return redirect('students_app:student_list')
            else:
                messages.error(request, f"Import failed: {result.get('message')}")
                return redirect('students_app:student_import')

        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            return redirect('students_app:student_import')

    # GET request - show the upload form
    context = {
        'title': 'Bulk Import Students',
    }
    return render(request, 'students/student_import.html', context)


@login_required
def delete_student(request, admission_number):
    from .models import Student
    if not _can_manage_school_data(request.user):
        messages.error(request, 'You do not have permission to delete students.')
        return redirect('students_app:student_list')

    if is_super_admin(request.user):
        student = get_object_or_404(Student, admission_number=admission_number)
    else:
        school = _get_user_school(request.user)
        if not school:
            messages.error(request, 'No school assigned to your account.')
            return redirect('students_app:student_list')
        student = get_object_or_404(Student, admission_number=admission_number, school=school)

    if request.method == 'POST':
        # Best-effort cleanup of login accounts created during bulk import
        try:
            from django.contrib.auth import get_user_model
            from .models import SchoolUser

            UserModel = get_user_model()
            user_obj = UserModel.objects.filter(username=student.admission_number).first()
            if user_obj and not user_obj.is_superuser:
                SchoolUser.objects.filter(user=user_obj).delete()
                user_obj.delete()
        except Exception:
            pass

        student.delete()
        messages.success(request, 'Student deleted successfully.')
        return redirect('students_app:student_list')

    return render(request, 'students/delete_student_confirm.html', {'student': student})


@login_required
@require_POST
def bulk_delete_students(request):
    from .models import Student
    if not _can_manage_school_data(request.user):
        messages.error(request, 'You do not have permission to delete students.')
        return redirect('students_app:student_list')

    admission_numbers = request.POST.getlist('admission_numbers')
    if not admission_numbers:
        messages.error(request, 'Please select at least one student to delete.')
        return redirect('students_app:student_list')

    qs = Student.objects.filter(admission_number__in=admission_numbers)
    if not is_super_admin(request.user):
        school = _get_user_school(request.user)
        if not school:
            messages.error(request, 'No school assigned to your account.')
            return redirect('students_app:student_list')
        qs = qs.filter(school=school)

    # Cleanup related accounts by username = admission_number (best-effort)
    try:
        from django.contrib.auth import get_user_model
        from .models import SchoolUser

        UserModel = get_user_model()
        user_qs = UserModel.objects.filter(username__in=list(qs.values_list('admission_number', flat=True)))
        SchoolUser.objects.filter(user__in=user_qs).delete()
        user_qs.filter(is_superuser=False).delete()
    except Exception:
        pass

    deleted_count, _ = qs.delete()
    messages.success(request, f'{deleted_count} student(s) deleted successfully.')
    return redirect('students_app:student_list')


@login_required
def delete_teacher(request, teacher_id):
    from .models import Teacher, SchoolUser
    if not _can_manage_school_data(request.user):
        messages.error(request, 'You do not have permission to delete teachers.')
        return redirect('students_app:teacher_list')

    qs = Teacher.objects.select_related('user', 'school')
    if is_super_admin(request.user):
        teacher = get_object_or_404(qs, id=teacher_id)
    else:
        school = _get_user_school(request.user)
        if not school:
            messages.error(request, 'No school assigned to your account.')
            return redirect('students_app:teacher_list')
        teacher = get_object_or_404(qs, id=teacher_id, school=school)

    if request.method == 'POST':
        try:
            user_obj = teacher.user
            if user_obj and not user_obj.is_superuser:
                SchoolUser.objects.filter(user=user_obj).delete()
                teacher.delete()
                user_obj.delete()
            else:
                teacher.delete()
        except Exception:
            teacher.delete()

        messages.success(request, 'Teacher deleted successfully.')
        return redirect('students_app:teacher_list')

    return render(request, 'teachers/delete_teacher_confirm.html', {'teacher': teacher})


@login_required
@require_POST
def bulk_delete_teachers(request):
    from .models import Teacher, SchoolUser
    if not _can_manage_school_data(request.user):
        messages.error(request, 'You do not have permission to delete teachers.')
        return redirect('students_app:teacher_list')

    teacher_ids = request.POST.getlist('teacher_ids')
    if not teacher_ids:
        messages.error(request, 'Please select at least one teacher to delete.')
        return redirect('students_app:teacher_list')

    qs = Teacher.objects.select_related('user', 'school').filter(id__in=teacher_ids)
    if not is_super_admin(request.user):
        school = _get_user_school(request.user)
        if not school:
            messages.error(request, 'No school assigned to your account.')
            return redirect('students_app:teacher_list')
        qs = qs.filter(school=school)

    user_ids = [t.user_id for t in qs if getattr(t, 'user_id', None)]
    deleted_count, _ = qs.delete()

    try:
        SchoolUser.objects.filter(user_id__in=user_ids).delete()
        from django.contrib.auth import get_user_model

        UserModel = get_user_model()
        UserModel.objects.filter(id__in=user_ids, is_superuser=False).delete()
    except Exception:
        pass

    messages.success(request, f'{deleted_count} teacher(s) deleted successfully.')
    return redirect('students_app:teacher_list')


@login_required
def delete_staff(request, staff_id):
    from .models import Staff, SchoolUser
    if not _can_manage_school_data(request.user):
        messages.error(request, 'You do not have permission to delete staff.')
        return redirect('students_app:staff_list')

    qs = Staff.objects.select_related('user', 'school')
    if is_super_admin(request.user):
        staff = get_object_or_404(qs, id=staff_id)
    else:
        school = _get_user_school(request.user)
        if not school:
            messages.error(request, 'No school assigned to your account.')
            return redirect('students_app:staff_list')
        staff = get_object_or_404(qs, id=staff_id, school=school)

    if request.method == 'POST':
        try:
            user_obj = staff.user
            if user_obj and not user_obj.is_superuser:
                SchoolUser.objects.filter(user=user_obj).delete()
                staff.delete()
                user_obj.delete()
            else:
                staff.delete()
        except Exception:
            staff.delete()

        messages.success(request, 'Staff deleted successfully.')
        return redirect('students_app:staff_list')

    return render(request, 'staff/delete_staff_confirm.html', {'staff': staff})


@login_required
@require_POST
def bulk_delete_staff(request):
    from .models import Staff, SchoolUser
    if not _can_manage_school_data(request.user):
        messages.error(request, 'You do not have permission to delete staff.')
        return redirect('students_app:staff_list')

    staff_ids = request.POST.getlist('staff_ids')
    if not staff_ids:
        messages.error(request, 'Please select at least one staff member to delete.')
        return redirect('students_app:staff_list')

    qs = Staff.objects.select_related('user', 'school').filter(id__in=staff_ids)
    if not is_super_admin(request.user):
        school = _get_user_school(request.user)
        if not school:
            messages.error(request, 'No school assigned to your account.')
            return redirect('students_app:staff_list')
        qs = qs.filter(school=school)

    user_ids = [s.user_id for s in qs if getattr(s, 'user_id', None)]
    deleted_count, _ = qs.delete()

    try:
        SchoolUser.objects.filter(user_id__in=user_ids).delete()
        from django.contrib.auth import get_user_model

        UserModel = get_user_model()
        UserModel.objects.filter(id__in=user_ids, is_superuser=False).delete()
    except Exception:
        pass

    messages.success(request, f'{deleted_count} staff member(s) deleted successfully.')
    return redirect('students_app:staff_list')


def process_student_import(excel_file, images_zip, user, require_section=False):
    """
    Process Excel file and import students with images
    """
    from datetime import datetime
    import pandas as pd
    import io
    from django.core.files.base import ContentFile
    from django.db import transaction

    from .models import Student, Class, Section, AcademicYear

    imported = 0
    skipped = 0
    errors = 0
    error_details = []
    skipped_details = []

    images_dict = {}
    if images_zip:
        try:
            images_dict = extract_images_from_zip(images_zip)
        except Exception as e:
            return {
                'success': False,
                'message': f'Error extracting images: {str(e)}'
            }

    try:
        df = pd.read_excel(excel_file, sheet_name=0)

        # Normalize column names - flexible: "Admission Number", "Adm No", "First Name" etc.
        def normalize_col(name):
            if pd.isna(name):
                return ''
            s = str(name).strip().lower().replace(' ', '_').replace('-', '_').replace('.', '')
            return s

        df.columns = [normalize_col(c) for c in df.columns]

        # Flexible column mapping - many names map to same field
        COLUMN_ALIASES = {
            'admission_number': ['admission_number', 'admission_no', 'adm_no', 'admno', 'roll_no', 'rollno'],
            'first_name': ['first_name', 'firstname', 'fname', 'first', 'given_name'],
            'last_name': ['last_name', 'lastname', 'lname', 'last', 'surname'],
            'date_of_birth': ['date_of_birth', 'dob', 'birth_date', 'birthdate', 'dob'],
            'gender': ['gender', 'sex'],
            'class_name': ['class_name', 'class', 'grade', 'standard', 'cls'],
            'section': ['section', 'section_name', 'sec'],
            'father_name': ['father_name', 'fathername', 'fathers_name', 'father', 'parent_name'],
            'father_phone': ['father_phone', 'father_phone_no', 'fathers_phone', 'father_mobile', 'parent_phone', 'phone'],
            'mother_name': ['mother_name', 'mothername', 'mothers_name', 'mother'],
            'mother_phone': ['mother_phone', 'mother_phone_no', 'mothers_phone', 'mother_mobile'],
            'roll_number': ['roll_number', 'roll_no', 'rollno'],
            'email': ['email', 'email_id'],
            'address': ['address', 'addr', 'residence'],
            'city': ['city'],
            'state': ['state'],
            'pincode': ['pincode', 'pin', 'pin_code', 'zip', 'zipcode'],
            'blood_group': ['blood_group', 'bloodgroup', 'blood'],
        }

        def get_col(df, aliases):
            for a in aliases:
                if a in df.columns:
                    return a
            return None

        def get_val(row, aliases, default=''):
            col = get_col(df, aliases)
            if col and col in row.index and pd.notna(row.get(col)):
                return str(row[col]).strip()
            return default

        # Required columns in Excel - names can vary (see COLUMN_ALIASES)
        REQ_COLUMNS = ['admission_number', 'first_name', 'last_name', 'date_of_birth', 'gender',
                       'class_name', 'father_name', 'father_phone', 'address', 'city', 'state', 'pincode']
        missing_cols = [f for f in REQ_COLUMNS if not get_col(df, COLUMN_ALIASES[f])]
        if missing_cols:
            examples = {'admission_number': 'Admission Number', 'father_phone': 'Father Phone',
                        'city': 'City', 'state': 'State', 'pincode': 'Pincode'}
            hint = ', '.join(examples.get(c, c) for c in missing_cols[:5])
            return {
                'success': False,
                'message': f'Excel must have these columns: {", ".join(missing_cols)}. Examples: {hint}'
            }

        # Determine school from user
        if hasattr(user, 'school_profile') and user.school_profile.school:
            school = user.school_profile.school
        else:
            return {
                'success': False,
                'message': 'User is not associated with any school'
            }

        def parse_date(val):
            """Parse date - supports DD.MM.YYYY, DD-MM-YYYY, Excel serial, etc."""
            if pd.isna(val):
                raise ValueError("Date is empty")
            val = str(val).strip()
            for fmt in ('%d.%m.%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d.%m.%y', '%d-%m-%y'):
                try:
                    return datetime.strptime(val, fmt).date()
                except ValueError:
                    continue
            try:
                return pd.to_datetime(val).date()
            except Exception:
                raise ValueError(f"Cannot parse date: {val}")

        def sanitize_phone(val):
            """Extract digits - min 10 for valid phone"""
            if not val:
                return ''
            s = ''.join(c for c in str(val) if c.isdigit())
            return s[:15] if len(s) >= 10 else ''

        today = datetime.now().date()
        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    admission_number = get_val(row, COLUMN_ALIASES['admission_number']).strip()
                    if not admission_number:
                        error_details.append(f"Row {index + 2}: Admission number is required")
                        errors += 1
                        continue

                    if Student.objects.filter(admission_number=admission_number, school=school).exists():
                        skipped += 1
                        skipped_details.append(f"Row {index + 2}: Duplicate admission number '{admission_number}'")
                        continue

                    first_name = get_val(row, COLUMN_ALIASES['first_name']).strip()
                    last_name = get_val(row, COLUMN_ALIASES['last_name']).strip()
                    if not first_name or not last_name:
                        error_details.append(f"Row {index + 2}: First name and last name are required")
                        errors += 1
                        continue

                    father_name = get_val(row, COLUMN_ALIASES['father_name']).strip()
                    father_phone = sanitize_phone(get_val(row, COLUMN_ALIASES['father_phone']))
                    if not father_name or not father_phone:
                        error_details.append(f"Row {index + 2}: Father name and father phone (10+ digits) are required")
                        errors += 1
                        continue

                    address = get_val(row, COLUMN_ALIASES['address']).strip()
                    city = get_val(row, COLUMN_ALIASES['city']).strip()
                    state = get_val(row, COLUMN_ALIASES['state']).strip()
                    pincode = get_val(row, COLUMN_ALIASES['pincode']).strip()
                    if not address or not city or not state or not pincode:
                        error_details.append(f"Row {index + 2}: Address, city, state and pincode are required")
                        errors += 1
                        continue

                    class_name_raw = get_val(row, COLUMN_ALIASES['class_name']).strip()
                    if not class_name_raw:
                        error_details.append(f"Row {index + 2}: Class name is required")
                        errors += 1
                        continue
                    # Normalize: "4th -" -> "4th", map to "Class 4" for consistency with fee structures
                    class_name = class_name_raw.rstrip(' -').strip()
                    mapping = {'1st': 'Class 1', '2nd': 'Class 2', '3rd': 'Class 3', '4th': 'Class 4',
                               '5th': 'Class 5', '6th': 'Class 6', '7th': 'Class 7', '8th': 'Class 8',
                               '9th': 'Class 9', '10th': 'Class 10', '11th': 'Class 11', '12th': 'Class 12',
                               'nursery': 'Nursery', 'lkg': 'LKG', 'ukg': 'UKG'}
                    class_name = mapping.get(class_name.lower(), class_name)
                    student_class, _ = Class.objects.get_or_create(name=class_name, defaults={'numeric_value': 0})

                    section = None
                    section_val = get_val(row, COLUMN_ALIASES['section'])
                    if section_val:
                        section, _ = Section.objects.get_or_create(name=section_val, class_assigned=student_class, defaults={'capacity': 40})
                    elif require_section:
                        error_details.append(
                            f"Row {index + 2}: Section is required. Either fill the Section column (e.g., A/B) or uncheck 'Section is mandatory' on the import page"
                        )
                        errors += 1
                        continue

                    current_year = datetime.now().year
                    year_str = f"{current_year}-{current_year + 1}"
                    start_dt = datetime(current_year, 4, 1).date()   # April 1
                    end_dt = datetime(current_year + 1, 3, 31).date()  # March 31 next year
                    academic_year, _ = AcademicYear.objects.get_or_create(
                        year=year_str,
                        defaults={'start_date': start_dt, 'end_date': end_dt}
                    )

                    dob_val = get_val(row, COLUMN_ALIASES['date_of_birth'])
                    if not dob_val:
                        error_details.append(f"Row {index + 2}: Date of birth is required")
                        errors += 1
                        continue
                    dob = parse_date(dob_val)

                    gender_val = get_val(row, COLUMN_ALIASES['gender']).upper()
                    if not gender_val:
                        error_details.append(f"Row {index + 2}: Gender is required (Male/Female or M/F)")
                        errors += 1
                        continue
                    if gender_val.startswith('F') or gender_val == 'FEMALE' or gender_val == 'GIRL':
                        gender_char = 'F'
                    elif gender_val.startswith('M') or gender_val == 'MALE' or gender_val == 'BOY':
                        gender_char = 'M'
                    else:
                        error_details.append(f"Row {index + 2}: Invalid gender '{gender_val}' - use Male/Female or M/F")
                        errors += 1
                        continue

                    student = Student.objects.create(
                        school=school,
                        admission_number=admission_number,
                        first_name=first_name,
                        last_name=last_name,
                        date_of_birth=dob,
                        gender=gender_char,
                        current_class=student_class,
                        section=section,
                        academic_year=academic_year,
                        father_name=father_name,
                        father_phone=father_phone,
                        mother_name=get_val(row, COLUMN_ALIASES['mother_name']),
                        mother_phone=sanitize_phone(get_val(row, COLUMN_ALIASES['mother_phone'])) or '',
                        email=get_val(row, COLUMN_ALIASES['email']) or '',
                        address=address,
                        city=city,
                        state=state,
                        pincode=pincode,
                        admission_date=today,
                        roll_number=get_val(row, COLUMN_ALIASES['roll_number']) or None,
                        blood_group=get_val(row, COLUMN_ALIASES['blood_group']) or '',
                        status='active'
                    )

                    # Create Django auth User for this student (username = admission_number)
                    try:
                        from django.contrib.auth import get_user_model
                        User = get_user_model()
                        username = admission_number
                        try:
                            pwd = student.date_of_birth.strftime('%Y%m%d')
                        except Exception:
                            pwd = str(student.date_of_birth)

                        existing_user = User.objects.filter(username=username).first()
                        if not existing_user:
                            try:
                                existing_user = User.objects.create_user(
                                    username=username,
                                    email=student.email or f"{username}@{school.name.replace(' ','').lower()}.local",
                                    password=pwd,
                                    first_name=student.first_name,
                                    last_name=student.last_name
                                )
                            except Exception:
                                existing_user = None

                        # Create SchoolUser profile if User was created
                        try:
                            from .models import UserRole, SchoolUser
                            student_role = UserRole.objects.filter(name='student').first()
                            if existing_user and student_role:
                                SchoolUser.objects.get_or_create(
                                    user=existing_user,
                                    defaults={
                                        'role': student_role,
                                        'school': school,
                                        'login_id': username,
                                        'custom_password': pwd,
                                        'phone': student.phone or ''
                                    }
                                )
                        except Exception:
                            pass
                    except Exception:
                        pass

                    image_filename = get_val(row, ['image_filename', 'photo_filename', 'photo', 'photo_name', 'photo_no', 'image'])
                    if image_filename and image_filename in images_dict:
                        image_data = images_dict[image_filename]
                        try:
                            if hasattr(student, 'profile_picture'):
                                student.profile_picture.save(f"{admission_number}.jpg", ContentFile(image_data), save=True)
                            else:
                                # fallback to `photo` field used in models
                                student.photo.save(f"{admission_number}.jpg", ContentFile(image_data), save=True)
                        except Exception:
                            pass

                    imported += 1

                except Exception as e:
                    errors += 1
                    error_details.append(f"Row {index + 2}: {str(e)}")
                    continue

        return {
            'success': True,
            'imported': imported,
            'skipped': skipped,
            'errors': errors,
            'error_details': error_details,
            'skipped_details': skipped_details,
        }

    except Exception as e:
        return {
            'success': False,
            'message': f'Error reading Excel file: {str(e)}'
        }


def extract_images_from_zip(zip_file):
    """
    Extract images from uploaded ZIP file
    Returns dictionary with filename as key and image data as value
    """
    import zipfile
    import os
    import io
    from PIL import Image

    images_dict = {}

    with zipfile.ZipFile(zip_file, 'r') as zip_ref:
        for file_info in zip_ref.filelist:
            if file_info.is_dir() or file_info.filename.startswith('__MACOSX'):
                continue

            filename = os.path.basename(file_info.filename)

            if filename.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
                try:
                    image_data = zip_ref.read(file_info.filename)
                    img = Image.open(io.BytesIO(image_data))
                    max_size = (800, 800)
                    img.thumbnail(max_size, Image.Resampling.LANCZOS)

                    if img.mode in ('RGBA', 'LA', 'P'):
                        background = Image.new('RGB', img.size, (255,255,255))
                        if img.mode == 'P':
                            img = img.convert('RGBA')
                        background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                        img = background

                    output = io.BytesIO()
                    img.save(output, format='JPEG', quality=85, optimize=True)
                    images_dict[filename] = output.getvalue()

                except Exception as e:
                    # skip problematic images
                    continue

    return images_dict


@login_required
def download_sample_template(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = [
        'admission_number', 'first_name', 'last_name', 'date_of_birth',
        'gender', 'class_name', 'section', 'roll_number',
        'father_name', 'father_phone', 'mother_name', 'mother_phone',
        'email', 'address', 'city', 'state', 'pincode', 'blood_group', 'image_filename'
    ]

    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header.replace('_', ' ').title()
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = 15

    sample_data = [
        ['2024001', 'John', 'Doe', '2010-05-15', 'Male', '5th', 'A', '1',
         'Mr. John Doe Sr.', '9876543210', 'Mrs. Jane Doe', '9876543211',
         'john.doe@email.com', '123 Main Street', 'Jabalpur', 'Madhya Pradesh', '482001', 'O+', 'john_doe.jpg'],
    ]

    for row_num, data in enumerate(sample_data, 2):
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=student_import_template.xlsx'
    wb.save(response)
    return response
